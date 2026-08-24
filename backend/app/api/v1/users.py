"""用户管理接口：CRUD / 数据范围过滤 / 导入导出 / 当前用户信息。"""
import csv
import io
import uuid

from fastapi import APIRouter, Depends, Query, Request, UploadFile
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_client_ip, get_current_user, get_user_agent, require_role
from app.core.exceptions import AppError, ERR_CONFLICT, ERR_NOT_FOUND, ERR_VALIDATION, ok_response
from app.core.security import hash_password
from app.db.session import get_db
from app.models import (
    AIConversation,
    Alert,
    AuditReport,
    Channel,
    ClientErrorReport,
    Department,
    Device,
    FileRecord,
    IPAllocation,
    LeaveRequest,
    Message,
    NetworkDiscovery,
    OperationLog,
    RefreshToken,
    Role,
    SandboxSession,
    ScanAuthorization,
    ScanReport,
    ScoreRecord,
    TrainingAgent,
    TrainingProgress,
    User,
    UserBadge,
)
from app.schemas.common import Page
from app.schemas.user import UserCreate, UserOut, UserUpdate
from app.services.audit_log import record
from app.services.data_scope import apply_data_scope

router = APIRouter(prefix="/users", tags=["用户管理"])

_IMPORT_COLUMNS = ["username", "real_name", "email", "phone", "employee_no", "department", "role", "position"]


_ADMIN_ROLE = "admin"
_ACTIVE_STATUSES = ("active", "on_leave", "business_trip")  # 在职/休假/外勤（离职归档与禁用除外）


async def _count_ref(session: AsyncSession, model: type, *conditions) -> int:
    """统计某表对指定用户的外键引用行数（支持同表 or_ 组合条件）。"""
    return (await session.execute(select(func.count()).select_from(model).where(*conditions))).scalar_one()


async def _validate_user_refs(session: AsyncSession, role_id: int | None = None, department_id: int | None = None) -> None:
    """外键存在性校验：role_id / department_id 引用不存在时返回 404（避免 FK IntegrityError）。"""
    if role_id is not None and not await session.get(Role, role_id):
        raise AppError(code=ERR_NOT_FOUND, message="角色不存在")
    if department_id is not None and not await session.get(Department, department_id):
        raise AppError(code=ERR_NOT_FOUND, message="部门不存在")


async def _count_active_admins(session: AsyncSession, exclude_id: int | None = None) -> int:
    """统计有效（在职/休假）管理员数量，可排除指定用户。"""
    query = (
        select(func.count(User.id))
        .join(Role, User.role_id == Role.id)
        .where(Role.code == _ADMIN_ROLE, User.status.in_(_ACTIVE_STATUSES))
    )
    if exclude_id:
        query = query.where(User.id != exclude_id)
    return (await session.execute(query)).scalar_one()


async def _role_code_of(session: AsyncSession, role_id: int | None) -> str | None:
    if not role_id:
        return None
    role = await session.get(Role, role_id)
    return role.code if role else None


def _to_out(user: User, roles: dict[int, Role], departments: dict[int, Department]) -> UserOut:
    role = roles.get(user.role_id) if user.role_id else None
    dept = departments.get(user.department_id) if user.department_id else None
    return UserOut(
        id=user.id,
        username=user.username,
        email=user.email,
        phone=user.phone,
        real_name=user.real_name,
        employee_no=user.employee_no,
        department_id=user.department_id,
        department_name=dept.name if dept else None,
        role_id=user.role_id,
        role=role.code if role else None,
        role_name=role.name if role else None,
        position=user.position,
        security_level=user.security_level,
        status=user.status,
        created_at=user.created_at,
        last_login_at=user.last_login_at,
        permissions=role.permissions if role else [],
    )


@router.get("/me")
async def get_me(
    session: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    dept = await session.get(Department, current.department_id) if current.department_id else None
    return ok_response(
        data=_to_out(current, {current.role_id: current._role} if current._role else {}, {dept.id: dept} if dept else {})
    )


@router.get("")
async def list_users(
    request: Request,
    keyword: str | None = None,
    department_id: int | None = None,
    status: str | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin", "manager"])),
):
    query = select(User)
    query = apply_data_scope(query, current, User)
    if keyword:
        like = f"%{keyword}%"
        query = query.where(or_(User.username.ilike(like), User.real_name.ilike(like), User.employee_no.ilike(like)))
    if department_id:
        query = query.where(User.department_id == department_id)
    if status:
        query = query.where(User.status == status)

    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (
        await session.execute(
            query.order_by(User.id).offset((page - 1) * size).limit(size)
        )
    ).scalars().all()

    role_ids = {u.role_id for u in rows if u.role_id}
    dept_ids = {u.department_id for u in rows if u.department_id}
    roles = {r.id: r for r in (await session.execute(select(Role).where(Role.id.in_(role_ids)))).scalars()} if role_ids else {}
    depts = {d.id: d for d in (await session.execute(select(Department).where(Department.id.in_(dept_ids)))).scalars()} if dept_ids else {}

    return ok_response(data=Page(items=[_to_out(u, roles, depts) for u in rows], total=total, page=page, size=size))


@router.post("")
async def create_user(
    data: UserCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin"])),
):
    await _validate_user_refs(session, data.role_id, data.department_id)
    exists = (await session.execute(select(User).where(User.username == data.username))).scalar_one_or_none()
    if exists:
        raise AppError(code=ERR_CONFLICT, message="用户名已存在")
    user = User(
        username=data.username,
        password_hash=hash_password(data.password),
        email=data.email,
        phone=data.phone,
        real_name=data.real_name,
        employee_no=data.employee_no,
        department_id=data.department_id,
        role_id=data.role_id,
        position=data.position,
        security_level=data.security_level,
        status=data.status,
    )
    session.add(user)
    await session.flush()
    await record(
        session, current, "user:create", target_type="user", target_id=str(user.id),
        detail={"username": user.username}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data=_to_out(user, {}, {}))


@router.put("/{user_id}")
async def update_user(
    user_id: int,
    data: UserUpdate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin"])),
):
    user = await session.get(User, user_id)
    if not user:
        raise AppError(code=ERR_NOT_FOUND, message="用户不存在")
    fields = data.model_dump(exclude_unset=True)
    await _validate_user_refs(session, fields.get("role_id"), fields.get("department_id"))

    # 自我保护：不能修改自己的角色，不能禁用/归档自己
    if user.id == current.id:
        if "role_id" in fields and fields["role_id"] != user.role_id:
            raise AppError(code=ERR_CONFLICT, message="不能修改自己的角色")
        if fields.get("status") in ("disabled", "archived"):
            raise AppError(code=ERR_CONFLICT, message="不能禁用或归档自己的账号")

    # 最后管理员保护：降级/禁用/归档管理员前必须保留至少一名有效管理员
    new_role_id = fields.get("role_id", user.role_id)
    new_status = fields.get("status", user.status)
    old_role_code = await _role_code_of(session, user.role_id)
    new_role_code = await _role_code_of(session, new_role_id)
    if old_role_code == _ADMIN_ROLE and new_role_code != _ADMIN_ROLE:
        if await _count_active_admins(session, exclude_id=user.id) < 1:
            raise AppError(code=ERR_CONFLICT, message="系统必须保留至少一名有效管理员")
    if (old_role_code == _ADMIN_ROLE or new_role_code == _ADMIN_ROLE) and new_status not in _ACTIVE_STATUSES:
        if await _count_active_admins(session, exclude_id=user.id) < 1:
            raise AppError(code=ERR_CONFLICT, message="系统必须保留至少一名有效管理员")

    if "password" in fields:
        fields["password_hash"] = hash_password(fields.pop("password"))
    for k, v in fields.items():
        setattr(user, k, v)
    await record(
        session, current, "user:update", target_type="user", target_id=str(user_id),
        detail={"fields": list(data.model_dump(exclude_unset=True).keys())},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    role = await session.get(Role, user.role_id) if user.role_id else None
    dept = await session.get(Department, user.department_id) if user.department_id else None
    return ok_response(data=_to_out(user, {user.role_id: role} if role else {}, {dept.id: dept} if dept else {}))


@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin"])),
):
    user = await session.get(User, user_id)
    if not user:
        raise AppError(code=ERR_NOT_FOUND, message="用户不存在")
    if user.id == current.id:
        raise AppError(code=ERR_CONFLICT, message="不能删除自己")
    if await _role_code_of(session, user.role_id) == _ADMIN_ROLE:
        if await _count_active_admins(session, exclude_id=user.id) < 1:
            raise AppError(code=ERR_CONFLICT, message="不能删除最后一名管理员")
    # 引用计数（对齐 departments.py 的 409+明细模式）：用户被业务数据引用时禁止物理删除，
    # 提示先解除引用；审计表只追加不可删除（RULE 保护），有审计记录的用户走归档保留审计链。
    uid = user_id
    refs = {
        "devices": await _count_ref(session, Device, Device.owner_id == uid),
        "ip_allocations": await _count_ref(session, IPAllocation, IPAllocation.allocated_to == uid),
        "alerts": await _count_ref(session, Alert, Alert.acknowledged_by == uid),
        "scan_reports": await _count_ref(session, ScanReport, or_(ScanReport.generated_by == uid, ScanReport.approved_by == uid)),
        "discoveries": await _count_ref(session, NetworkDiscovery, NetworkDiscovery.created_by == uid),
        "channels": await _count_ref(session, Channel, Channel.creator_id == uid),
        "messages": await _count_ref(session, Message, Message.sender_id == uid),
        "ai_conversations": await _count_ref(session, AIConversation, AIConversation.user_id == uid),
        "files": await _count_ref(session, FileRecord, FileRecord.user_id == uid),
        "leave_requests": await _count_ref(session, LeaveRequest, or_(LeaveRequest.user_id == uid, LeaveRequest.approver_id == uid)),
        "training": (
            await _count_ref(session, TrainingProgress, TrainingProgress.user_id == uid)
            + await _count_ref(session, SandboxSession, SandboxSession.user_id == uid)
            + await _count_ref(session, ScoreRecord, ScoreRecord.user_id == uid)
        ),
        "badges": await _count_ref(session, UserBadge, UserBadge.user_id == uid),
        "error_reports": await _count_ref(session, ClientErrorReport, ClientErrorReport.user_id == uid),
        "scan_authorizations": await _count_ref(session, ScanAuthorization, ScanAuthorization.approved_by == uid),
        "training_agents": await _count_ref(session, TrainingAgent, TrainingAgent.created_by == uid),
        "audit_reports": await _count_ref(session, AuditReport, AuditReport.generated_by == uid),
    }
    has_logs = (await session.execute(select(func.count()).select_from(OperationLog).where(OperationLog.user_id == uid))).scalar_one() > 0
    if not has_logs and sum(refs.values()) > 0:
        _REF_LABELS = {
            "devices": "设备", "ip_allocations": "IP 分配", "alerts": "告警", "scan_reports": "扫描报告",
            "discoveries": "网络发现", "channels": "频道", "messages": "消息", "ai_conversations": "AI 会话",
            "files": "上传文件", "leave_requests": "休假申请", "training": "训练记录", "badges": "徽章",
            "error_reports": "错误上报", "scan_authorizations": "扫描授权",
            "training_agents": "训练代理", "audit_reports": "审计报告",
        }
        detail = "、".join(f"{_REF_LABELS[k]} {v} 条" for k, v in refs.items() if v)
        raise AppError(
            code=ERR_CONFLICT,
            message=f"该用户存在业务引用（{detail}），请先解除引用后再删除",
            data=refs,
        )
    action = "archived" if has_logs else "deleted"
    if has_logs:
        user.status = "archived"
        user.password_hash = hash_password(f"__archived_{user.id}_{uuid.uuid4().hex}")  # 口令立即失效
    await record(
        session, current, "user:delete", target_type="user", target_id=str(user_id),
        detail={"username": user.username, "action": action, "refs": refs},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    if has_logs:
        await session.commit()
        return ok_response(data={"action": action})
    # refresh token 是纯会话记录（非业务数据）：物理删除用户前先清除，等效注销其全部会话
    await session.execute(delete(RefreshToken).where(RefreshToken.user_id == user_id))
    await session.delete(user)
    await session.commit()
    return ok_response(data={"action": action})


@router.post("/import")
async def import_users(
    file: UploadFile,
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin"])),
):
    """批量导入用户（CSV 或 XLSX）。表头：username,real_name,email,phone,employee_no,department,role,position"""
    raw = await file.read()
    filename = file.filename or ""
    rows: list[dict] = []
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw), read_only=True)
            ws = wb.active
            header = [c.value for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2):
                rows.append({header[i]: (row[i].value if i < len(row) else None) for i in range(len(header))})
        except Exception:
            raise AppError(code=ERR_VALIDATION, message="Excel 文件解析失败，请确认是有效的 .xlsx 文件")
    else:
        try:
            text = raw.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        except Exception:
            raise AppError(code=ERR_VALIDATION, message="CSV 文件解析失败，请使用 UTF-8 编码")

    roles = {r.code: r for r in (await session.execute(select(Role))).scalars()}
    depts = {d.name: d for d in (await session.execute(select(Department))).scalars()}

    created, errors = 0, []
    for idx, row in enumerate(rows, start=2):
        username = (row.get("username") or "").strip()
        if not username:
            errors.append({"row": idx, "error": "缺少 username"})
            continue
        if (await session.execute(select(User).where(User.username == username))).scalar_one_or_none():
            errors.append({"row": idx, "error": f"用户名 {username} 已存在"})
            continue
        employee_no = (row.get("employee_no") or "").strip()
        if employee_no and (await session.execute(select(User).where(User.employee_no == employee_no))).scalar_one_or_none():
            errors.append({"row": idx, "error": f"工号 {employee_no} 已存在"})
            continue
        role = roles.get((row.get("role") or "").strip())
        dept = depts.get((row.get("department") or "").strip())
        user = User(
            username=username,
            password_hash=hash_password(row.get("password") or "Bt@123456"),
            real_name=row.get("real_name"),
            email=row.get("email"),
            phone=row.get("phone"),
            employee_no=row.get("employee_no"),
            role_id=role.id if role else None,
            department_id=dept.id if dept else None,
            position=row.get("position"),
        )
        session.add(user)
        created += 1
    await record(session, current, "user:import", detail={"created": created, "failed": len(errors)}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request))
    await session.commit()
    return ok_response(data={"created": created, "failed": errors})


@router.get("/export")
async def export_users(
    request: Request,
    session: AsyncSession = Depends(get_db),
    current: User = Depends(require_role(["admin", "manager"])),
):
    query = apply_data_scope(select(User), current, User)
    users = (await session.execute(query.order_by(User.id))).scalars().all()
    await record(session, current, "user:export", ip_address=await get_client_ip(request), user_agent=await get_user_agent(request))
    await session.commit()

    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM，Excel 直接打开中文不乱码（与导入 utf-8-sig 对齐）
    writer = csv.DictWriter(buf, fieldnames=_IMPORT_COLUMNS + ["password"])
    writer.writeheader()
    for u in users:
        row = {c: (getattr(u, c, "") or "") for c in _IMPORT_COLUMNS}
        # department/role 是关系对象，导出名称与角色编码（与导入模板一致，可回导）
        row["department"] = u.department.name if u.department else ""
        row["role"] = u.role.code if u.role else ""
        writer.writerow(row | {"password": ""})
    from fastapi.responses import StreamingResponse

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=users.csv"},
    )
