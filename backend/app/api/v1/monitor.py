"""监控子系统 API：设备 / IPAM / 告警 / 漏洞扫描（真实 nmap）。"""
import asyncio
import datetime as dt
import io
import ipaddress
import re

from fastapi import APIRouter, Depends, Header, Query, Request, UploadFile
from sqlalchemy import delete, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.dependencies import get_client_ip, get_current_user, get_user_agent, require_permission, require_role
from app.core.exceptions import AppError, ERR_CONFLICT, ERR_FORBIDDEN, ERR_NOT_FOUND, ERR_VALIDATION, ok_response
from app.db.session import get_db
from app.models import (
    Alert, ClientErrorReport, Department, Device, DevicePatrol, IPAllocation, IPSubnet, NetworkDiscovery,
    OperationLog, Role, ScanAuthorization, ScanReport, User,
)
from app.schemas.common import Page
from app.schemas.monitor import (
    AlertCreate, AllocationCreate, AllocationUpdate, ClientErrorReportIn, DeviceCreate, DeviceUpdate,
    DiscoveryCreate, DiscoveryRegister, ScanAuthCreate, ScanCreate, SubnetCreate, SubnetUpdate,
)
from app.services.audit_log import record
from app.services.data_scope import apply_data_scope, apply_device_data_scope
from app.services.notify import notify_alert_task
from app.services.scan_policy import is_internal_ip, is_internal_network
from app.services.scanner import (
    ScanOptions,
    _options_from_dict,
    _options_to_dict,
    _validate_port_range,
    cancel_scan,
    launch_discovery,
    launch_scan,
)

router = APIRouter(tags=["监控中心"])

_ALERT_STATUSES = ("open", "acknowledged", "resolved")

_MAC_RE = re.compile(r"^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$")


async def _validate_device_refs(session: AsyncSession, department_id=None, owner_id=None, mac_address=None) -> None:
    """设备外键与格式校验：department_id / owner_id 引用不存在返回 404；MAC 地址格式非法返回 400。"""
    if department_id is not None and not await session.get(Department, department_id):
        raise AppError(code=ERR_NOT_FOUND, message="部门不存在")
    if owner_id is not None and not await session.get(User, owner_id):
        raise AppError(code=ERR_NOT_FOUND, message="负责人不存在")
    if mac_address and not _MAC_RE.match(mac_address):
        raise AppError(code=ERR_VALIDATION, message="MAC 地址格式不正确（如 AA:BB:CC:DD:EE:FF）")


async def _device_in_scope(session: AsyncSession, user: User, device_id: int) -> bool:
    """设备是否在用户数据范围内（写操作门禁，与 get_device 一致）。"""
    scoped = apply_data_scope(select(Device.id).where(Device.id == device_id), user, Device)
    return bool((await session.execute(scoped)).scalar_one_or_none())


async def _subnet_in_scope(session: AsyncSession, user: User, subnet_id: int | None) -> bool:
    """子网是否在用户数据范围内（IPAM 写操作门禁，对齐 update_allocation）。"""
    if not subnet_id:
        return True
    scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == subnet_id), user, IPSubnet)
    return bool((await session.execute(scoped)).scalar_one_or_none())


# ---------- 设备 ----------
def _device_out(d: Device, depts: dict, owners: dict) -> dict:
    return {
        "id": d.id, "name": d.name, "ip_address": d.ip_address, "mac_address": d.mac_address,
        "device_type": d.device_type, "manufacturer": d.manufacturer, "model": d.model,
        "location": d.location, "department_id": d.department_id,
        "department_name": depts.get(d.department_id) if depts else None,
        "owner_id": d.owner_id, "owner_name": owners.get(d.owner_id) if owners else None,
        "snmp_community": d.snmp_community, "status": d.status,
        "last_seen_at": d.last_seen_at, "offline_since": d.offline_since, "created_at": d.created_at,
    }


@router.get("/monitor/devices")
async def list_devices(
    keyword: str | None = None,
    status: str | None = None,
    device_type: str | None = None,
    department_id: int | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    query = select(Device)
    query = apply_data_scope(query, user, Device)
    if keyword:
        like = f"%{keyword}%"
        query = query.where(or_(Device.name.ilike(like), Device.ip_address.ilike(like), Device.location.ilike(like)))
    if status:
        query = query.where(Device.status == status)
    if device_type:
        query = query.where(Device.device_type == device_type)
    if department_id:
        query = query.where(Device.department_id == department_id)

    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(Device.id).offset((page - 1) * size).limit(size))).scalars().all()
    depts = {d.id: d.name for d in (await session.execute(select(Department))).scalars()}
    owners = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_({d.owner_id for d in rows if d.owner_id})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[_device_out(d, depts, owners) for d in rows], total=total, page=page, size=size))


@router.post("/monitor/devices")
async def create_device(
    data: DeviceCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:manage")),
):
    await _validate_device_refs(session, data.department_id, data.owner_id, data.mac_address)
    exists = (await session.execute(select(Device).where(Device.ip_address == data.ip_address))).scalar_one_or_none()
    if exists:
        raise AppError(code=ERR_CONFLICT, message="该 IP 已被设备占用")
    payload = data.model_dump()
    if payload.get("department_id") is None:
        payload["department_id"] = user.department_id  # 归属创建者部门，保证 dept 数据范围自可见
    d = Device(**payload)
    session.add(d)
    await session.flush()
    await record(
        session, user, "monitor:device:create", target_type="device", target_id=str(d.id),
        detail={"name": d.name, "ip": d.ip_address}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data=_device_out(d, {}, {}))


@router.get("/monitor/devices/{device_id:int}")
async def get_device(
    device_id: int,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    d = await session.get(Device, device_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    # 数据范围：与 list_devices 一致（dept/self 角色不得越权读他部门设备）
    scoped = apply_data_scope(select(Device.id).where(Device.id == device_id), user, Device)
    if not (await session.execute(scoped)).scalar_one_or_none():
        raise AppError(code=ERR_FORBIDDEN, message="无权查看该设备")
    depts = {d.department_id: (await session.get(Department, d.department_id)).name} if d.department_id else {}
    owner = await session.get(User, d.owner_id) if d.owner_id else None
    return ok_response(data=_device_out(d, depts, {d.owner_id: (owner.real_name or owner.username)} if owner else {}))


@router.put("/monitor/devices/{device_id}")
async def update_device(
    device_id: int,
    data: DeviceUpdate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:manage")),
):
    d = await session.get(Device, device_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    if not await _device_in_scope(session, user, device_id):
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该设备")
    payload = data.model_dump(exclude_unset=True)
    await _validate_device_refs(session, payload.get("department_id"), payload.get("owner_id"), payload.get("mac_address"))
    if "status" in payload and payload["status"] == "archived" and d.status != "archived":
        # 归档设备
        payload["status"] = "archived"
    for k, v in payload.items():
        setattr(d, k, v)
    await record(
        session, user, "monitor:device:update", target_type="device", target_id=str(device_id),
        detail={"changes": list(payload.keys())}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response()


@router.delete("/monitor/devices/{device_id}")
async def delete_device(
    device_id: int,
    request: Request,
    reason: str | None = Query(None, max_length=200),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:manage")),
):
    d = await session.get(Device, device_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    if not await _device_in_scope(session, user, device_id):
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该设备")
    refs = {
        "alerts": (await session.execute(select(func.count()).select_from(Alert).where(Alert.device_id == device_id))).scalar_one(),
        "allocations": (await session.execute(select(func.count()).select_from(IPAllocation).where(IPAllocation.device_id == device_id))).scalar_one(),
        "scan_reports": (await session.execute(select(func.count()).select_from(ScanReport).where(ScanReport.device_id == device_id))).scalar_one(),
    }
    if sum(refs.values()) > 0:
        # 有引用关联（告警/IP 分配/扫描报告）→ 归档保留数据链
        d.status = "archived"
        message = "设备存在关联数据（告警/IP 分配/扫描报告），已归档保留"
    else:
        await session.delete(d)
        message = "设备已删除"
    await record(
        session, user, "monitor:device:delete", target_type="device", target_id=str(device_id),
        detail={"archived": sum(refs.values()) > 0, "refs": refs, "reason": reason},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"message": message})


_DEVICE_EXPORT_COLUMNS = ["名称", "IP地址", "MAC地址", "设备类型", "厂商", "型号", "位置", "部门", "负责人", "状态", "最近在线", "创建时间"]
# 导入表头归一：兼容导出的中文表头与模板英文表头
_DEVICE_HEADER_ALIAS = {
    "名称": "name", "IP地址": "ip_address", "IP 地址": "ip_address",
    "MAC地址": "mac_address", "MAC 地址": "mac_address",
    "设备类型": "device_type", "厂商": "manufacturer", "型号": "model",
    "位置": "location", "部门": "department", "负责人": "owner", "状态": "status",
}


def _norm_device_row(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        key = str(k).strip() if k is not None else ""
        out[_DEVICE_HEADER_ALIAS.get(key, key)] = v
    return out


@router.get("/monitor/devices/export")
async def export_devices(
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    """导出设备清单为 XLSX（与 list_devices 相同的数据范围约束）。"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    rows = (await session.execute(apply_data_scope(select(Device), user, Device).order_by(Device.id))).scalars().all()
    depts = {d.id: d.name for d in (await session.execute(select(Department))).scalars()}
    owners = {u.id: (u.real_name or u.username) for u in (await session.execute(
        select(User).where(User.id.in_({d.owner_id for d in rows if d.owner_id}))
    )).scalars()} if rows else {}
    await record(
        session, user, "monitor:device:export", target_type="device", detail={"count": len(rows)},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()

    wb = Workbook()
    ws = wb.active
    ws.title = "设备清单"
    ws.append(_DEVICE_EXPORT_COLUMNS)
    for d in rows:
        ws.append([
            d.name, str(d.ip_address) if d.ip_address else "", d.mac_address or "", d.device_type or "",
            d.manufacturer or "", d.model or "", d.location or "",
            depts.get(d.department_id) or "", owners.get(d.owner_id) or "",
            d.status,
            d.last_seen_at.strftime("%Y-%m-%d %H:%M") if d.last_seen_at else "",
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=devices.xlsx"},
    )


_DEVICE_IMPORT_COLUMNS = ["name", "ip_address", "mac_address", "device_type", "manufacturer", "model", "location", "department", "status"]


@router.post("/monitor/devices/import")
async def import_devices(
    file: UploadFile,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:manage")),
):
    """批量导入设备（CSV 或 XLSX）。表头：name,ip_address,mac_address,device_type,manufacturer,model,location,department,status"""
    filename = file.filename or ""
    raw = await file.read()
    rows: list[dict] = []
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True)
        except Exception:
            raise AppError(code=ERR_VALIDATION, message="Excel 文件无法解析，请使用模板")
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows())]
        for r in ws.iter_rows(min_row=2):
            rows.append(_norm_device_row({header[i]: (r[i].value if i < len(r) else None) for i in range(len(header))}))
    elif filename.lower().endswith(".csv"):
        import csv

        rows = [_norm_device_row(r) for r in csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))]
    else:
        raise AppError(code=ERR_VALIDATION, message="仅支持 .xlsx / .csv 文件")

    depts = {d.name: d for d in (await session.execute(select(Department))).scalars()}
    created, errors = 0, []
    for idx, row in enumerate(rows, start=2):
        values = row
        name = (values.get("name") or "").strip()
        ip = (values.get("ip_address") or "").strip()
        if not name or not ip:
            errors.append({"row": idx, "error": "缺少 name / ip_address"})
            continue
        try:
            ipaddress.ip_address(ip)
        except ValueError:
            errors.append({"row": idx, "error": f"IP {ip} 格式不正确"})
            continue
        if (await session.execute(select(Device).where(Device.ip_address == ip))).scalar_one_or_none():
            errors.append({"row": idx, "error": f"IP {ip} 已存在"})
            continue
        status = (values.get("status") or "active").strip()
        if status not in ("active", "offline", "maintenance", "archived"):
            errors.append({"row": idx, "error": f"状态 {status} 非法"})
            continue
        mac = (values.get("mac_address") or "").strip()
        if mac and not _MAC_RE.match(mac):
            errors.append({"row": idx, "error": f"MAC {mac} 格式不正确"})
            continue
        dept_name = (values.get("department") or "").strip()
        if dept_name and dept_name not in depts:
            errors.append({"row": idx, "error": f"部门 {dept_name} 不存在"})
            continue
        session.add(Device(
            name=name, ip_address=ip,
            mac_address=mac or None,
            device_type=(values.get("device_type") or "").strip() or None,
            manufacturer=(values.get("manufacturer") or "").strip() or None,
            model=(values.get("model") or "").strip() or None,
            location=(values.get("location") or "").strip() or None,
            department_id=depts[dept_name].id if dept_name in depts else None,
            status=status,
        ))
        created += 1
    await record(
        session, user, "monitor:device:import", detail={"created": created, "failed": len(errors)},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"created": created, "failed": errors})


@router.post("/monitor/devices/{device_id}/ping")
async def ping_device(
    device_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:manage")),
):
    # 探测会改写 status/last_seen_at，属写动作，需 manage 权限（原 view 已收紧，审计用只读角色不再能探测）
    d = await session.get(Device, device_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    if not await _device_in_scope(session, user, device_id):
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该设备")
    if d.status == "archived":
        raise AppError(code=ERR_VALIDATION, message="设备已归档")
    d.last_seen_at = dt.datetime.now(dt.timezone.utc)
    d.status = "active"
    await record(
        session, user, "monitor:device:ping", target_type="device", target_id=str(device_id),
        detail={"ip": str(d.ip_address)},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"last_seen_at": d.last_seen_at})


# ---------- IPAM ----------


async def _recycle_expired_leases(session: AsyncSession) -> int:
    """惰性回收过期 DHCP 租约：物理删除，使地址可立即重新分配。

    仅处理 allocation_type == dhcp；static / reserved 不做自动回收，避免服务器固定 IP 意外释放。
    释放留痕由审计日志（ipam:alloc:release / create）承担，故此处真删而非软删——
    ip_allocations 对 ip_address 有物理唯一约束，软删会导致地址永不复用。
    """
    now = dt.datetime.now(dt.timezone.utc)
    result = await session.execute(
        delete(IPAllocation)
        .where(
            IPAllocation.allocation_type == "dhcp",
            IPAllocation.expires_at.is_not(None),
            IPAllocation.expires_at < now,
            IPAllocation.is_active.is_(True),
        )
    )
    count = result.rowcount or 0
    if count:
        await session.commit()
    return count


@router.post("/monitor/allocations/recycle")
async def recycle_expired_leases_now(
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    """手动回收所有过期 DHCP 租约（定时任务之外的手动兜底入口，返回回收条数）。"""
    count = await _recycle_expired_leases(session)
    await record(
        session, user, "ipam:lease:recycle", detail={"recycled": count},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"recycled": count})


def _validate_reserved_ranges(ranges: list[str] | None, net) -> list:
    """校验保留地址段：每个元素须为合法 CIDR 且落在该子网内，返回规范化后的列表。"""
    if not ranges:
        return []
    reserved = []
    for r in ranges:
        try:
            # strict=False：容忍主机位非零的输入（如 10.0.10.10/28），归一化为网络地址存储
            rn = ipaddress.ip_network(r, strict=False)
        except ValueError:
            raise AppError(code=ERR_VALIDATION, message=f"保留段 {r} 格式不正确，应为 CIDR（如 10.0.10.100/28）")
        if not rn.subnet_of(net):
            raise AppError(code=ERR_VALIDATION, message=f"保留段 {r} 不在子网 {net} 内")
        reserved.append(str(rn))
    return reserved


def _subnet_usage(subnet: IPSubnet) -> int:
    try:
        net = ipaddress.ip_network(subnet.network)
        return net.num_addresses
    except Exception:
        return 0


@router.get("/monitor/subnets")
async def list_subnets(
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    # 先回收过期 DHCP 租约，使用量统计反映真实占用
    await _recycle_expired_leases(session)
    # 部门数据范围：与设备列表一致，manager/admin 全量，analyst 仅本部门
    query = select(IPSubnet).where(IPSubnet.is_active.is_(True))
    query = apply_data_scope(query, user, IPSubnet)
    subnets = (await session.execute(query.order_by(IPSubnet.id))).scalars().all()
    depts = {d.id: d.name for d in (await session.execute(select(Department))).scalars()}
    used_map = {
        s_id: c
        for s_id, c in (
            await session.execute(
                select(IPAllocation.subnet_id, func.count()).where(IPAllocation.subnet_id.in_([s.id for s in subnets])).group_by(IPAllocation.subnet_id)
            )
        ).all()
    }
    return ok_response(data=[
        {
            "id": s.id, "name": s.name, "network": s.network, "gateway": s.gateway, "dns_servers": s.dns_servers,
            "vlan_id": s.vlan_id, "department_id": s.department_id, "department_name": depts.get(s.department_id),
            "reserved_ranges": s.reserved_ranges or [],
            "capacity": _subnet_usage(s), "used": used_map.get(s.id, 0),
        }
        for s in subnets
    ])


@router.post("/monitor/subnets")
async def create_subnet(
    data: SubnetCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    try:
        net = ipaddress.ip_network(data.network)
    except ValueError:
        raise AppError(code=ERR_VALIDATION, message="网段格式不正确")
    # 仅内网网段可入台账（IPAM 是内网资产管理，登记公网网段无意义且扩大扫描面）
    if not is_internal_network(net):
        raise AppError(code=ERR_VALIDATION, message="子网必须是内网网段")
    # 同网段不可重复登记
    dup = (await session.execute(
        select(IPSubnet).where(IPSubnet.network == str(net), IPSubnet.is_active.is_(True))
    )).scalar_one_or_none()
    if dup:
        raise AppError(code=ERR_CONFLICT, message="该网段已登记")
    # 重叠/嵌套检测：与任意已登记网段有交叉即拒绝，防止分配冲突
    existing = (await session.execute(select(IPSubnet).where(IPSubnet.is_active.is_(True)))).scalars().all()
    for other in existing:
        try:
            other_net = ipaddress.ip_network(other.network)
        except ValueError:
            continue
        if net.overlaps(other_net):
            raise AppError(
                code=ERR_CONFLICT,
                message=f"网段 {str(other.network)}（{other.name}）与当前网段重叠或包含，请调整规划",
            )
    if data.department_id is not None and not await session.get(Department, data.department_id):
        raise AppError(code=ERR_NOT_FOUND, message="部门不存在")
    payload = data.model_dump()
    # 校验保留地址段，并规范化为 CIDR 列表
    payload["reserved_ranges"] = _validate_reserved_ranges(data.reserved_ranges, net)
    # 显式传入网关时校验 IP 格式（对齐 update_subnet）
    if payload.get("gateway"):
        try:
            ipaddress.ip_address(payload["gateway"])
        except ValueError:
            raise AppError(code=ERR_VALIDATION, message="网关格式不正确")
    # 网关留空 → 自动取子网第一个可用地址（如 10.0.30.1）
    if not payload.get("gateway") and net.version == 4:
        try:
            payload["gateway"] = str(next(net.hosts()))
        except StopIteration:
            payload["gateway"] = None
    s = IPSubnet(**payload)
    session.add(s)
    await session.flush()
    await record(
        session, user, "ipam:subnet:create", target_type="subnet", target_id=str(s.id),
        detail={"name": s.name, "network": str(s.network), "gateway": s.gateway},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"id": s.id, "network": s.network, "gateway": s.gateway})


@router.delete("/monitor/subnets/{subnet_id}")
async def delete_subnet(
    subnet_id: int,
    request: Request,
    reason: str | None = Query(None, max_length=200),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    s = await session.get(IPSubnet, subnet_id)
    if not s or not s.is_active:
        raise AppError(code=ERR_NOT_FOUND, message="子网不存在")
    active_alloc = (await session.execute(
        select(func.count()).select_from(IPAllocation).where(
            IPAllocation.subnet_id == subnet_id, IPAllocation.is_active.is_(True)
        )
    )).scalar_one()
    if active_alloc:
        raise AppError(code=ERR_CONFLICT, message=f"子网下仍有 {active_alloc} 条地址分配，请先释放")
    # 软删：is_active=False，保留历史（分配记录 FK 依赖子网，不可物理删除）
    s.is_active = False
    await record(
        session, user, "ipam:subnet:delete", target_type="subnet", target_id=str(subnet_id),
        detail={"name": s.name, "network": str(s.network), "reason": reason},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"message": "子网已删除"})


@router.put("/monitor/subnets/{subnet_id}")
async def update_subnet(
    subnet_id: int,
    data: SubnetUpdate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    s = await session.get(IPSubnet, subnet_id)
    if not s or not s.is_active:
        raise AppError(code=ERR_NOT_FOUND, message="子网不存在")
    net = ipaddress.ip_network(s.network)
    changes: dict = {}
    if data.name is not None and data.name != s.name:
        changes["name"] = data.name
        s.name = data.name
    if data.gateway is not None:
        try:
            gw = ipaddress.ip_address(data.gateway)
        except ValueError:
            raise AppError(code=ERR_VALIDATION, message="网关格式不正确")
        if gw not in net:
            raise AppError(code=ERR_VALIDATION, message="网关不在该网段内")
        if str(gw) != s.gateway:
            changes["gateway"] = str(gw)
            s.gateway = str(gw)
    if data.dns_servers is not None and data.dns_servers != (s.dns_servers or []):
        changes["dns_servers"] = data.dns_servers
        s.dns_servers = data.dns_servers
    if data.vlan_id is not None and data.vlan_id != s.vlan_id:
        changes["vlan_id"] = data.vlan_id
        s.vlan_id = data.vlan_id
    if data.department_id is not None and data.department_id != s.department_id:
        if not await session.get(Department, data.department_id):
            raise AppError(code=ERR_NOT_FOUND, message="部门不存在")
        changes["department_id"] = data.department_id
        s.department_id = data.department_id
    if data.reserved_ranges is not None:
        reserved = _validate_reserved_ranges(data.reserved_ranges, net)
        if reserved != (s.reserved_ranges or []):
            changes["reserved_ranges"] = reserved
            s.reserved_ranges = reserved
    if not changes:
        return ok_response(data={"message": "无变更"})
    await record(
        session, user, "ipam:subnet:update", target_type="subnet", target_id=str(subnet_id),
        detail={"name": s.name, "network": str(s.network), "changes": changes},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"id": s.id, "network": s.network, "gateway": s.gateway, "changes": changes})


@router.get("/monitor/subnets/{subnet_id}/usage")
async def subnet_usage(
    subnet_id: int,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    """子网地址使用明细：供前端渲染 IP 使用率热图（无分页）。"""
    s = await session.get(IPSubnet, subnet_id)
    if not s or not s.is_active:
        raise AppError(code=ERR_NOT_FOUND, message="子网不存在")
    scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == s.id), user, IPSubnet)
    if not (await session.execute(scoped)).scalar_one_or_none():
        raise AppError(code=ERR_FORBIDDEN, message="无权查看该子网")
    allocs = (await session.execute(
        select(IPAllocation).where(IPAllocation.subnet_id == s.id, IPAllocation.is_active.is_(True))
    )).scalars().all()
    return ok_response(data={
        "id": s.id, "name": s.name, "network": s.network, "gateway": s.gateway,
        "vlan_id": s.vlan_id, "reserved_ranges": s.reserved_ranges or [],
        "capacity": _subnet_usage(s), "used": len(allocs),
        "allocations": [
            {
                "ip": str(a.ip_address), "allocation_type": a.allocation_type,
                "purpose": a.purpose, "expires_at": a.expires_at,
            }
            for a in allocs
        ],
    })


@router.get("/monitor/allocations")
async def list_allocations(
    subnet_id: int | None = None,
    keyword: str | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    # 先回收过期 DHCP 租约，列表不含已到期记录
    await _recycle_expired_leases(session)
    # IPAM 数据范围：先求当前用户可见子网集合，再过滤分配记录
    scope_query = apply_data_scope(select(IPSubnet.id).where(IPSubnet.is_active.is_(True)), user, IPSubnet)
    visible_subnet_ids = set((await session.execute(scope_query)).scalars())
    query = select(IPAllocation).where(
        IPAllocation.is_active.is_(True),
        IPAllocation.subnet_id.in_(visible_subnet_ids),
    )
    if subnet_id:
        query = query.where(IPAllocation.subnet_id == subnet_id)
    if keyword:
        like = f"%{keyword}%"
        query = query.where(or_(IPAllocation.ip_address.ilike(like), IPAllocation.purpose.ilike(like)))
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(IPAllocation.id.desc()).offset((page - 1) * size).limit(size))).scalars().all()
    subnets = {s.id: s for s in (await session.execute(select(IPSubnet).where(IPSubnet.id.in_({a.subnet_id for a in rows if a.subnet_id})))).scalars()} if rows else {}
    users = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_({a.allocated_to for a in rows if a.allocated_to})))).scalars()} if rows else {}
    devices = {d.id: d.name for d in (await session.execute(select(Device).where(Device.id.in_({a.device_id for a in rows if a.device_id})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[
        {
            "id": a.id, "subnet_id": a.subnet_id, "subnet_name": subnets[a.subnet_id].name if a.subnet_id in subnets else None,
            "ip_address": a.ip_address, "device_id": a.device_id, "device_name": devices.get(a.device_id),
            "allocated_to": a.allocated_to, "allocated_to_name": users.get(a.allocated_to),
            "allocation_type": a.allocation_type, "purpose": a.purpose, "expires_at": a.expires_at,
        }
        for a in rows
    ], total=total, page=page, size=size))


def _next_free_ip(session: AsyncSession, subnet: IPSubnet, existing: set[str]) -> str | None:
    net = ipaddress.ip_network(subnet.network)
    # gateway 列（INET）从 DB 读出是 IPv4Address 对象，统一转 str 再比较
    gw = str(subnet.gateway) if subnet.gateway else None
    # 保留地址段：自动分配跳过（手动指定不受限）
    reserved = []
    for r in (subnet.reserved_ranges or []):
        try:
            reserved.append(ipaddress.ip_network(r))
        except ValueError:
            continue
    hosts = net.hosts() if net.version == 4 else list(net.subnets(prefixlen_diff=1))  # IPv6 简化处理
    for ip in hosts:
        s = str(ip)
        if gw and s == gw:
            continue
        if any(ipaddress.ip_address(s) in rn for rn in reserved):
            continue
        if s not in existing:
            return s
    return None


@router.post("/monitor/allocations")
async def create_allocation(
    data: AllocationCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    if not data.subnet_id:
        raise AppError(code=ERR_VALIDATION, message="必须指定子网")
    subnet = await session.get(IPSubnet, data.subnet_id)
    if not subnet or not subnet.is_active:
        raise AppError(code=ERR_NOT_FOUND, message="子网不存在")
    # IPAM 数据范围：仅允许操作当前用户可见范围内的子网
    scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == subnet.id), user, IPSubnet)
    if not (await session.execute(scoped)).scalar_one_or_none():
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该子网")
    if data.allocated_to is not None and not await session.get(User, data.allocated_to):
        raise AppError(code=ERR_NOT_FOUND, message="分配用户不存在")
    if data.device_id is not None and not await session.get(Device, data.device_id):
        raise AppError(code=ERR_NOT_FOUND, message="绑定设备不存在")
    # 先回收过期 DHCP 租约，确保自动分配拿到已到期地址
    await _recycle_expired_leases(session)

    # ip_address 为 INET 列，读回是 IPv4Address 对象，统一转 str 才能与自动分配比较
    taken = set(
        str(ip)
        for ip in (await session.execute(
            select(IPAllocation.ip_address).where(IPAllocation.subnet_id == subnet.id, IPAllocation.is_active.is_(True))
        )).scalars()
    )
    if data.ip_address:
        try:
            ip = ipaddress.ip_address(data.ip_address)
        except ValueError:
            raise AppError(code=ERR_VALIDATION, message="IP 格式不正确")
        if ip not in ipaddress.ip_network(subnet.network):
            raise AppError(code=ERR_VALIDATION, message="IP 不在该子网范围内")
        if data.ip_address in taken:
            raise AppError(code=ERR_CONFLICT, message="该 IP 已分配")
        addr = data.ip_address
    else:
        addr = _next_free_ip(session, subnet, taken)
        if not addr:
            raise AppError(code=ERR_VALIDATION, message="子网地址已用尽")
    # 全局唯一约束兜底
    dup = (await session.execute(select(IPAllocation).where(IPAllocation.ip_address == addr, IPAllocation.is_active.is_(True)))).scalar_one_or_none()
    if dup:
        raise AppError(code=ERR_CONFLICT, message="该 IP 已在其他子网占用")

    a = IPAllocation(
        subnet_id=subnet.id, ip_address=addr, allocated_to=data.allocated_to, device_id=data.device_id,
        allocation_type=data.allocation_type, purpose=data.purpose, expires_at=data.expires_at,
    )
    session.add(a)
    await session.flush()
    await record(
        session, user, "ipam:alloc:create", target_type="allocation", target_id=str(a.id),
        detail={"ip": addr, "subnet": str(subnet.network)}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"id": a.id, "ip_address": addr, "subnet_id": subnet.id, "subnet_name": subnet.name})


@router.delete("/monitor/allocations/{allocation_id}")
async def release_allocation(
    allocation_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    a = await session.get(IPAllocation, allocation_id)
    if not a:
        raise AppError(code=ERR_NOT_FOUND, message="分配记录不存在")
    if not await _subnet_in_scope(session, user, a.subnet_id):
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该子网的分配")
    released_ip = str(a.ip_address)
    # 真删：ip_allocations 对 ip_address 有物理唯一约束，软删会导致该 IP 永不复用
    await session.delete(a)
    await record(
        session, user, "ipam:alloc:release", target_type="allocation", target_id=str(allocation_id),
        detail={"ip": released_ip}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response()


@router.get("/monitor/allocations/history")
async def allocation_history(
    ip: str,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    """按 IP 查询该地址的分配/释放/修改审计轨迹（时间倒序）。"""
    logs = (await session.execute(
        select(OperationLog)
        .where(
            OperationLog.action.in_(["ipam:alloc:create", "ipam:alloc:release", "ipam:alloc:update"]),
            OperationLog.detail["ip"].astext == ip,
        )
        .order_by(OperationLog.id.desc())
        .limit(50)
    )).scalars().all()
    return ok_response(data=[
        {
            "id": log.id, "action": log.action, "username": log.username, "role_code": log.role_code,
            "detail": log.detail, "created_at": log.created_at,
        }
        for log in logs
    ])


@router.put("/monitor/allocations/{allocation_id}")
async def update_allocation(
    allocation_id: int,
    data: AllocationUpdate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    a = await session.get(IPAllocation, allocation_id)
    if not a:
        raise AppError(code=ERR_NOT_FOUND, message="分配记录不存在")
    # IPAM 数据范围：分配所属子网需在用户可见范围内
    if a.subnet_id:
        subnet = await session.get(IPSubnet, a.subnet_id)
        if subnet:
            scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == subnet.id), user, IPSubnet)
            if not (await session.execute(scoped)).scalar_one_or_none():
                raise AppError(code=ERR_FORBIDDEN, message="无权操作该子网的分配")
    changes: dict = {}
    for field in ("purpose", "allocated_to", "device_id", "allocation_type", "expires_at"):
        val = getattr(data, field)
        if val is not None and getattr(a, field) != val:
            changes[field] = val
            setattr(a, field, val)
    if not changes:
        return ok_response(data={"message": "无变更"})
    await record(
        session, user, "ipam:alloc:update", target_type="allocation", target_id=str(allocation_id),
        detail={"ip": str(a.ip_address), "changes": changes},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"id": a.id, "ip_address": a.ip_address, "changes": changes})


# ---------- 告警 ----------
def _alert_out(a: Alert, devices: dict, users: dict) -> dict:
    return {
        "id": a.id, "device_id": a.device_id, "device_name": devices.get(a.device_id),
        "alert_type": a.alert_type, "severity": a.severity, "title": a.title, "description": a.description,
        "status": a.status, "acknowledged_by": a.acknowledged_by,
        "acknowledged_by_name": users.get(a.acknowledged_by),
        "resolved_at": a.resolved_at, "notified_at": a.notified_at, "created_at": a.created_at,
    }


@router.get("/monitor/alerts")
async def list_alerts(
    status: str | None = None,
    severity: str | None = None,
    device_id: int | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:alert:view")),
):
    query = select(Alert)
    query = apply_device_data_scope(query, user, Alert)
    if status:
        query = query.where(Alert.status == status)
    if severity:
        query = query.where(Alert.severity == severity)
    if device_id:
        query = query.where(Alert.device_id == device_id)
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(Alert.id.desc()).offset((page - 1) * size).limit(size))).scalars().all()
    devices = {d.id: d.name for d in (await session.execute(select(Device).where(Device.id.in_({a.device_id for a in rows if a.device_id})))).scalars()} if rows else {}
    users = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_({a.acknowledged_by for a in rows if a.acknowledged_by})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[_alert_out(a, devices, users) for a in rows], total=total, page=page, size=size))


@router.post("/monitor/alerts")
async def create_alert(
    data: AlertCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:alert:manage")),
):
    if data.device_id:
        d = await session.get(Device, data.device_id)
        if not d:
            raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    a = Alert(**data.model_dump())
    session.add(a)
    await session.flush()
    await record(
        session, user, "monitor:alert:create", target_type="alert", target_id=str(a.id),
        detail={"title": a.title, "severity": a.severity}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    # 外部通知（企业微信/钉钉/邮件）：后台任务发送，成功后回写 notified_at；失败静默
    asyncio.create_task(notify_alert_task(a.id, a.title, a.description or "", a.severity))
    return ok_response(data={"id": a.id, "status": a.status})


async def _alert_workflow(session, user, alert_id, action: str, request: Request):
    a = await session.get(Alert, alert_id)
    if not a:
        raise AppError(code=ERR_NOT_FOUND, message="告警不存在")
    if action == "acknowledge":
        if a.status in ("acknowledged", "resolved"):
            raise AppError(code=ERR_VALIDATION, message="告警已处理")
        a.status = "acknowledged"
        a.acknowledged_by = user.id
    elif action == "resolve":
        if a.status == "resolved":
            raise AppError(code=ERR_VALIDATION, message="告警已解决")
        a.status = "resolved"
        a.acknowledged_by = user.id
        a.resolved_at = dt.datetime.now(dt.timezone.utc)
    await record(
        session, user, f"monitor:alert:{action}", target_type="alert", target_id=str(alert_id),
        detail={"status": a.status}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response()


@router.post("/monitor/alerts/{alert_id}/acknowledge")
async def acknowledge_alert(
    alert_id: int, request: Request,
    session: AsyncSession = Depends(get_db), user: User = Depends(require_permission("monitor:alert:manage")),
):
    return await _alert_workflow(session, user, alert_id, "acknowledge", request)


@router.post("/monitor/alerts/{alert_id}/resolve")
async def resolve_alert(
    alert_id: int, request: Request,
    session: AsyncSession = Depends(get_db), user: User = Depends(require_permission("monitor:alert:manage")),
):
    return await _alert_workflow(session, user, alert_id, "resolve", request)


# ---------- 客户端错误上报（前端全局捕获，匿名允许） ----------
async def _optional_current_user(
    request: Request,
    session: AsyncSession = Depends(get_db),
    authorization: str | None = Header(default=None),
) -> User | None:
    """尝试识别当前用户；未认证/识别失败返回 None（匿名上报不拦截）。"""
    try:
        return await get_current_user(request, authorization=authorization, session=session)
    except Exception:
        return None


@router.post("/monitor/client-errors")
async def report_client_error(
    data: ClientErrorReportIn,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User | None = Depends(_optional_current_user),
):
    """前端运行时错误上报：落 client_error_reports，供无后端异常时定位页面故障。"""
    session.add(ClientErrorReport(
        user_id=user.id if user else None,
        url=data.url, message=data.message[:1000], stack=data.stack,
        user_agent=(request.headers.get("user-agent") or "")[:300],
    ))
    await session.commit()
    return ok_response()


# ---------- 扫描授权名单（内网硬限制 + 授权判定） ----------
async def _in_authorized_scope(
    session: AsyncSession, target: ipaddress.IPv4Address | ipaddress.IPv6Address | ipaddress.IPv4Network | ipaddress.IPv6Network
) -> bool:
    """目标（IP 或网段）是否在授权扫描范围内：active 子网台账 ∪ active 未过期授权名单。"""
    now = dt.datetime.now(dt.timezone.utc)
    subnets = (await session.execute(
        select(IPSubnet).where(IPSubnet.is_active.is_(True))
    )).scalars().all()
    auths = (await session.execute(
        select(ScanAuthorization).where(ScanAuthorization.status == "active")
    )).scalars().all()
    candidates = [ipaddress.ip_network(s.network) for s in subnets]
    for a in auths:
        if a.start_date and a.start_date > now:
            continue
        if a.end_date and a.end_date < now:
            continue
        candidates.append(ipaddress.ip_network(a.network))
    if isinstance(target, (ipaddress.IPv4Address, ipaddress.IPv6Address)):
        return any(target in c for c in candidates)
    return any(target.subnet_of(c) for c in candidates)


@router.get("/monitor/scan-auth")
async def list_scan_auth(
    status: str | None = None,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    """扫描授权名单：active/revoked 全量，含到期时间（供前端管理）。"""
    query = select(ScanAuthorization)
    if status:
        query = query.where(ScanAuthorization.status == status)
    rows = (await session.execute(query.order_by(ScanAuthorization.id.desc()))).scalars().all()
    approvers = {u.id: (u.real_name or u.username) for u in (
        await session.execute(select(User).where(User.id.in_({a.approved_by for a in rows if a.approved_by})))
    ).scalars()} if rows else {}
    now = dt.datetime.now(dt.timezone.utc)
    return ok_response(data=[{
        "id": a.id, "name": a.name, "network": a.network, "status": a.status,
        "start_date": a.start_date, "end_date": a.end_date, "note": a.note,
        "approved_by": a.approved_by, "approved_by_name": approvers.get(a.approved_by),
        "expired": bool(a.end_date and a.end_date < now),
        "created_at": a.created_at,
    } for a in rows])


@router.post("/monitor/scan-auth")
async def create_scan_auth(
    data: ScanAuthCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_role(["manager", "admin"])),
):
    """登记扫描授权网段：仅内网网段可授权，支持有效期。"""
    try:
        net = ipaddress.ip_network(data.network, strict=False)
    except ValueError:
        raise AppError(code=ERR_VALIDATION, message=f"网段格式不正确：{data.network}")
    if not is_internal_network(net):
        raise AppError(code=ERR_VALIDATION, message="扫描授权仅允许内网网段")
    if data.start_date and data.end_date and data.start_date > data.end_date:
        raise AppError(code=ERR_VALIDATION, message="生效时间不能晚于到期时间")
    a = ScanAuthorization(
        name=data.name, network=str(net), start_date=data.start_date, end_date=data.end_date,
        note=data.note, approved_by=user.id,
    )
    session.add(a)
    await session.flush()
    await record(
        session, user, "ipam:scan_auth:create", target_type="scan_authorization", target_id=str(a.id),
        detail={"name": a.name, "network": str(net), "end_date": a.end_date.isoformat() if a.end_date else None},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"id": a.id, "network": a.network})


@router.post("/monitor/scan-auth/{auth_id}/revoke")
async def revoke_scan_auth(
    auth_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_role(["manager", "admin"])),
):
    """吊销扫描授权：立即失效，不可恢复（重新登记即可）。"""
    a = await session.get(ScanAuthorization, auth_id)
    if not a:
        raise AppError(code=ERR_NOT_FOUND, message="授权记录不存在")
    if a.status == "revoked":
        raise AppError(code=ERR_CONFLICT, message="授权已吊销")
    a.status = "revoked"
    await record(
        session, user, "ipam:scan_auth:revoke", target_type="scan_authorization", target_id=str(auth_id),
        detail={"name": a.name, "network": str(a.network)},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(message="授权已吊销")


# ---------- 漏洞扫描（真实 nmap，异步） ----------
async def _validate_scan_target(session: AsyncSession, target_ip: str) -> None:
    """扫描目标限制：仅内网地址，且必须落在授权范围（登记子网 ∪ 授权名单）内。"""
    try:
        ip = ipaddress.ip_address(target_ip)
    except ValueError:
        raise AppError(code=ERR_VALIDATION, message="IP 格式不正确")
    if not is_internal_ip(ip):
        raise AppError(code=ERR_VALIDATION, message="扫描目标必须是内网地址")
    if not await _in_authorized_scope(session, ip):
        raise AppError(code=ERR_VALIDATION, message="目标不在已登记网段或扫描授权名单内")


@router.post("/monitor/scans")
async def create_scan(
    data: ScanCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    if data.device_id:
        d = await session.get(Device, data.device_id)
        if not d:
            raise AppError(code=ERR_NOT_FOUND, message="设备不存在")
    if data.ports and data.port_range:
        raise AppError(code=ERR_VALIDATION, message="端口数与端口范围二选一")
    if data.port_range:
        try:
            _validate_port_range(data.port_range)
        except ValueError as e:
            raise AppError(code=ERR_VALIDATION, message=str(e))
    await _validate_scan_target(session, data.target_ip)
    opts = ScanOptions(
        scan_type=data.scan_type or settings.NMAP_SCAN_TYPE,
        top_ports=data.ports,
        port_range=data.port_range,
        service_detection=settings.NMAP_VERSION_DETECT,
        nse=data.nse,
    )
    report = ScanReport(
        report_type=data.report_type, device_id=data.device_id, target_ip=data.target_ip,
        scan_data=None, risk_score=None,
        summary=f"目标 {data.target_ip} 扫描排队中…", status="pending_review",
        scan_status="pending", generated_by=user.id, scan_options=_options_to_dict(opts),
    )
    session.add(report)
    await session.flush()
    await record(
        session, user, "monitor:scan:create", target_type="scan_report", target_id=str(report.id),
        detail={"target": data.target_ip, "scan_type": opts.scan_type, "nse": opts.nse},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    launch_scan(report.id, data.target_ip, opts)  # 后台执行，立即返回
    return ok_response(data={"report_id": report.id, "scan_status": "pending"})


@router.get("/monitor/scans/reports")
async def list_scan_reports(
    status: str | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    query = select(ScanReport)
    query = apply_device_data_scope(query, user, ScanReport)
    if status:
        query = query.where(ScanReport.status == status)
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(ScanReport.id.desc()).offset((page - 1) * size).limit(size))).scalars().all()
    devices = {d.id: d.name for d in (await session.execute(select(Device).where(Device.id.in_({r.device_id for r in rows if r.device_id})))).scalars()} if rows else {}
    users = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_({r.generated_by for r in rows if r.generated_by} | {r.approved_by for r in rows if r.approved_by})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[
        {
            "id": r.id, "report_type": r.report_type, "device_id": r.device_id, "device_name": devices.get(r.device_id),
            "target_ip": r.target_ip, "risk_score": r.risk_score, "summary": r.summary, "status": r.status,
            "scan_status": r.scan_status, "error": r.error, "error_code": r.error_code,
            "scan_options": r.scan_options,
            "generated_by": r.generated_by, "generated_by_name": users.get(r.generated_by),
            "approved_by": r.approved_by, "approved_by_name": users.get(r.approved_by),
            "generated_at": r.generated_at,
        }
        for r in rows
    ], total=total, page=page, size=size))


@router.get("/monitor/scans/reports/{report_id}")
async def get_scan_report(
    report_id: int,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    r = await session.get(ScanReport, report_id)
    if not r:
        raise AppError(code=ERR_NOT_FOUND, message="报告不存在")
    # 数据范围：与 list_scan_reports 一致（按设备归属过滤）；创建者本人始终可看自己的报告
    # （无设备关联的目标 IP 扫描报告，dept/self 角色不能因无设备归属而失去对自己报告的访问）
    scoped = apply_device_data_scope(select(ScanReport.id).where(ScanReport.id == report_id), user, ScanReport)
    if r.generated_by != user.id and not (await session.execute(scoped)).scalar_one_or_none():
        raise AppError(code=ERR_FORBIDDEN, message="无权查看该报告")
    users = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_([r.generated_by, r.approved_by] if r.approved_by else [r.generated_by])))).scalars()} if (r.generated_by or r.approved_by) else {}
    device = await session.get(Device, r.device_id) if r.device_id else None
    return ok_response(data={
        "id": r.id, "report_type": r.report_type, "device_id": r.device_id, "device_name": device.name if device else None,
        "target_ip": r.target_ip, "scan_data": r.scan_data, "summary": r.summary, "risk_score": r.risk_score,
        "status": r.status, "scan_status": r.scan_status, "error": r.error, "error_code": r.error_code,
        "scan_options": r.scan_options,
        "generated_by_name": users.get(r.generated_by),
        "approved_by_name": users.get(r.approved_by), "generated_at": r.generated_at,
    })


@router.post("/monitor/scans/reports/{report_id}/cancel")
async def cancel_scan_api(
    report_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    """取消进行中的扫描任务（pending/running → failed/cancelled）。"""
    r = await session.get(ScanReport, report_id)
    if not r:
        raise AppError(code=ERR_NOT_FOUND, message="报告不存在")
    if r.scan_status not in ("pending", "running"):
        raise AppError(code=ERR_VALIDATION, message="当前状态不可取消")
    if not cancel_scan(report_id):
        # 进程内无活动任务（单 worker 极罕见）→ 幂等直接落失败
        r.scan_status, r.error, r.error_code = "failed", "扫描已被取消", "cancelled"
    await record(
        session, user, "monitor:scan:cancel", target_type="scan_report", target_id=str(report_id),
        detail={}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(message="已请求取消")


@router.post("/monitor/scans/reports/{report_id}/retry")
async def retry_scan_api(
    report_id: int,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:scan")),
):
    """重试失败的扫描（仅 failed 可重试；沿用原扫描选项，重置执行生命周期）。"""
    r = await session.get(ScanReport, report_id)
    if not r:
        raise AppError(code=ERR_NOT_FOUND, message="报告不存在")
    if r.scan_status != "failed":
        raise AppError(code=ERR_VALIDATION, message="仅失败任务可重试")
    opts = _options_from_dict(r.scan_options)
    r.scan_status, r.error, r.error_code = "pending", None, None
    r.scan_data, r.risk_score = None, None
    r.summary = f"目标 {r.target_ip} 重试排队中…"
    await record(
        session, user, "monitor:scan:retry", target_type="scan_report", target_id=str(report_id),
        detail={}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    launch_scan(report_id, str(r.target_ip), opts)  # INET 列读回的是 ipaddress 对象，str 归一化
    return ok_response(data={"report_id": report_id, "scan_status": "pending"})


@router.post("/monitor/scans/reports/{report_id}/review")
async def review_scan_report(
    report_id: int,
    request: Request,
    approve: bool = True,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_role(["manager", "admin"])),
):
    r = await session.get(ScanReport, report_id)
    if not r:
        raise AppError(code=ERR_NOT_FOUND, message="报告不存在")
    if r.scan_status != "completed":
        raise AppError(code=ERR_VALIDATION, message="扫描尚未完成，无法审核")
    if r.status != "pending_review":
        raise AppError(code=ERR_VALIDATION, message="报告已审核")
    r.status = "approved" if approve else "rejected"
    r.approved_by = user.id
    await record(
        session, user, "monitor:scan:review", target_type="scan_report", target_id=str(report_id),
        detail={"approve": approve}, ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"status": r.status})


# ---------- 网络发现（主机发现 + 终端登记） ----------
@router.post("/monitor/discover")
async def create_discovery(
    data: DiscoveryCreate,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    subnet = None
    if data.subnet_id:
        subnet = await session.get(IPSubnet, data.subnet_id)
        if not subnet or not subnet.is_active:
            raise AppError(code=ERR_NOT_FOUND, message="子网不存在")
        scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == subnet.id), user, IPSubnet)
        if not (await session.execute(scoped)).scalar_one_or_none():
            raise AppError(code=ERR_FORBIDDEN, message="无权扫描该子网")
    # 目标网段：优先用请求体 network，关联子网时可省略由子网网段兜底
    target = data.network or (str(subnet.network) if subnet else None)
    if not target:
        raise AppError(code=ERR_VALIDATION, message="请提供目标网段或关联子网")
    try:
        net = ipaddress.ip_network(target)
    except ValueError:
        raise AppError(code=ERR_VALIDATION, message=f"网段格式不正确：{target}")
    # 仅内网网段可发现（公网探测一票否决）
    if not is_internal_network(net):
        raise AppError(code=ERR_VALIDATION, message="发现目标必须是内网网段")
    # 防超大网段：主机发现对地址空间全扫描，/22（1024）为上限（资源保护优先于授权判定）
    if net.num_addresses > 1024:
        raise AppError(code=ERR_VALIDATION, message="网段过大，请选择 /22 或更小网段")
    # 授权校验：登记子网 ∪ 扫描授权名单；手动网段需先登记授权
    if not await _in_authorized_scope(session, net):
        raise AppError(code=ERR_VALIDATION, message="目标网段未在扫描授权范围内，请先在扫描授权名单中登记")
    d = NetworkDiscovery(subnet_id=subnet.id if subnet else None, network=str(net), scan_status="pending")
    session.add(d)
    await session.flush()
    await record(
        session, user, "ipam:discover:create", target_type="discovery", target_id=str(d.id),
        detail={"subnet": subnet.name if subnet else None, "network": str(net)},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    launch_discovery(d.id, str(net))  # 后台执行，立即返回
    return ok_response(data={"discovery_id": d.id, "scan_status": "pending"})


@router.get("/monitor/discover/{discovery_id}")
async def get_discovery(
    discovery_id: int,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    d = await session.get(NetworkDiscovery, discovery_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="发现记录不存在")
    # 数据范围：对齐 list_discoveries——关联子网的发现按子网归属过滤；
    # 手动网段（subnet_id IS NULL）记录仅全量权限角色（admin/manager）可读。
    role = getattr(user, "_role", None)
    full_scope = role is not None and (role.data_scope == "all" or role.code == "admin")
    if d.subnet_id:
        scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == d.subnet_id), user, IPSubnet)
        if not (await session.execute(scoped)).scalar_one_or_none():
            raise AppError(code=ERR_FORBIDDEN, message="无权查看该发现记录")
    elif not full_scope:
        raise AppError(code=ERR_FORBIDDEN, message="无权查看该发现记录")
    subnet = await session.get(IPSubnet, d.subnet_id) if d.subnet_id else None
    creator = await session.get(User, d.created_by) if d.created_by else None
    net = ipaddress.ip_network(str(d.network))
    # IPv4 用点分掩码（255.255.255.0），IPv6 无传统掩码则返回 /prefix
    netmask = str(net.netmask) if net.version == 4 else f"/{net.prefixlen}"
    return ok_response(data={
        "id": d.id, "subnet_id": d.subnet_id, "subnet_name": subnet.name if subnet else None,
        "network": d.network, "netmask": netmask, "scan_status": d.scan_status, "error": d.error,
        "hosts": d.hosts or [],
        "online_ips": d.online_ips or [], "unregistered_ips": d.unregistered_ips or [],
        "registered_ips": d.registered_ips or [], "offline_ips": d.offline_ips or [],
        "created_by_name": (creator.real_name or creator.username) if creator else None,
        "created_at": d.created_at, "completed_at": d.completed_at,
    })


@router.get("/monitor/discover")
async def list_discoveries(
    subnet_id: int | None = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    # 数据范围：先求可见子网集合。全量权限角色（manager/admin）可见全部记录（含手动网段
    # subnet_id IS NULL 的发现）；部门范围角色仅可见其部门子网关联的发现。
    scope_query = apply_data_scope(select(IPSubnet.id).where(IPSubnet.is_active.is_(True)), user, IPSubnet)
    visible_subnet_ids = set((await session.execute(scope_query)).scalars())
    all_subnet_ids = set((await session.execute(select(IPSubnet.id).where(IPSubnet.is_active.is_(True)))).scalars())
    query = select(NetworkDiscovery)
    if visible_subnet_ids != all_subnet_ids:
        query = query.where(NetworkDiscovery.subnet_id.in_(visible_subnet_ids))
    if subnet_id:
        query = query.where(NetworkDiscovery.subnet_id == subnet_id)
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(NetworkDiscovery.id.desc()).offset((page - 1) * size).limit(size))).scalars().all()
    subnets = {s.id: s for s in (await session.execute(select(IPSubnet).where(IPSubnet.id.in_({d.subnet_id for d in rows if d.subnet_id})))).scalars()} if rows else {}
    users = {u.id: (u.real_name or u.username) for u in (await session.execute(select(User).where(User.id.in_({d.created_by for d in rows if d.created_by})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[
        {
            "id": d.id, "subnet_id": d.subnet_id, "subnet_name": subnets[d.subnet_id].name if d.subnet_id in subnets else None,
            "network": d.network, "scan_status": d.scan_status, "error": d.error,
            "online_count": len(d.online_ips or []), "unregistered_count": len(d.unregistered_ips or []),
            "registered_count": len(d.registered_ips or []), "offline_count": len(d.offline_ips or []),
            "created_by_name": users.get(d.created_by), "created_at": d.created_at,
        }
        for d in rows
    ], total=total, page=page, size=size))


# ---------- 设备自动巡检（后台定时，状态刷新 + 结果追溯） ----------
@router.get("/monitor/patrols")
async def list_patrols(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("monitor:device:view")),
):
    # 数据范围：与 list_discoveries 同模式——先求可见子网集合；部门范围角色
    # 仅可见其部门子网关联的巡检记录（自然排除手动/无子网记录）。
    scope_query = apply_data_scope(select(IPSubnet.id).where(IPSubnet.is_active.is_(True)), user, IPSubnet)
    visible_subnet_ids = set((await session.execute(scope_query)).scalars())
    all_subnet_ids = set((await session.execute(select(IPSubnet.id).where(IPSubnet.is_active.is_(True)))).scalars())
    query = select(DevicePatrol)
    if visible_subnet_ids != all_subnet_ids:
        query = query.where(DevicePatrol.subnet_id.in_(visible_subnet_ids))
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    rows = (await session.execute(query.order_by(DevicePatrol.id.desc()).offset((page - 1) * size).limit(size))).scalars().all()
    subnets = {s.id: s for s in (await session.execute(select(IPSubnet).where(IPSubnet.id.in_({p.subnet_id for p in rows if p.subnet_id})))).scalars()} if rows else {}
    return ok_response(data=Page(items=[
        {
            "id": p.id, "subnet_id": p.subnet_id, "subnet_name": subnets[p.subnet_id].name if p.subnet_id in subnets else None,
            "network": p.network, "scan_status": p.scan_status, "error": p.error,
            "online_count": len(p.online_ips or []), "offline_count": len(p.offline_ips or []),
            "ghost_count": len(p.ghost_ips or []),
            "started_at": p.started_at, "completed_at": p.completed_at, "created_at": p.created_at,
        }
        for p in rows
    ], total=total, page=page, size=size))


@router.post("/monitor/discover/{discovery_id}/register")
async def register_discovery(
    discovery_id: int,
    data: DiscoveryRegister,
    request: Request,
    session: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("ipam:manage")),
):
    """把发现结果中勾选的幽灵设备登记为「终端设备 + DHCP 分配」（半自动确认，单事务原子）。

    MAC/厂商从本次发现的 hosts 元数据反查（防前端篡改）；子网归属优先复用
    关联/匹配子网，否则自动创建，使网段掩码固化进 IPAM 台账。
    """
    d = await session.get(NetworkDiscovery, discovery_id)
    if not d:
        raise AppError(code=ERR_NOT_FOUND, message="发现记录不存在")
    if d.scan_status != "completed":
        raise AppError(code=ERR_VALIDATION, message="发现尚未完成，无法登记")
    scoped = apply_data_scope(select(IPSubnet.id).where(IPSubnet.id == d.subnet_id), user, IPSubnet)
    if d.subnet_id and not (await session.execute(scoped)).scalar_one_or_none():
        raise AppError(code=ERR_FORBIDDEN, message="无权操作该子网")
    net = ipaddress.ip_network(str(d.network))
    host_by_ip = {h["ip"]: h for h in (d.hosts or [])}
    ips = list(dict.fromkeys(data.ips))  # 去重，保持输入顺序
    for ip_str in ips:
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            raise AppError(code=ERR_VALIDATION, message=f"IP 格式不正确：{ip_str}")
        if ip not in net:
            raise AppError(code=ERR_VALIDATION, message=f"IP 不在该子网范围内：{ip_str}")
        if (await session.execute(select(Device).where(Device.ip_address == ip_str))).scalar_one_or_none():
            raise AppError(code=ERR_CONFLICT, message=f"IP 已登记为设备：{ip_str}")
        dup = (await session.execute(
            select(IPAllocation).where(IPAllocation.ip_address == ip_str, IPAllocation.is_active.is_(True))
        )).scalar_one_or_none()
        if dup:
            raise AppError(code=ERR_CONFLICT, message=f"IP 已被登记：{ip_str}")
        mac = (host_by_ip.get(ip_str) or {}).get("mac")
        if mac:
            mac_owner = (await session.execute(select(Device).where(Device.mac_address == mac))).scalar_one_or_none()
            if mac_owner:
                raise AppError(code=ERR_CONFLICT, message=f"MAC 已登记为其他设备：{mac}")
    # 子网归属（掩码入台账）：关联子网 → 复用匹配子网 → 自动创建
    subnet = await session.get(IPSubnet, d.subnet_id) if d.subnet_id else None
    auto_subnet = False
    if subnet is None:
        subnet = (await session.execute(select(IPSubnet).where(IPSubnet.network == str(net)))).scalars().first()
    if subnet is None:
        subnet = IPSubnet(name=f"终端网段-{str(net)}", network=str(net))
        session.add(subnet)
        await session.flush()
        auto_subnet = True
    purpose = data.purpose or "网络发现登记"
    for ip_str in ips:
        meta = host_by_ip.get(ip_str) or {}
        dev = Device(
            name=f"终端-{ip_str}", ip_address=ip_str,
            mac_address=meta.get("mac"), manufacturer=meta.get("vendor"),
            department_id=subnet.department_id, status="active",
        )
        session.add(dev)
        await session.flush()
        session.add(IPAllocation(
            subnet_id=subnet.id, ip_address=ip_str, device_id=dev.id,
            allocation_type="dhcp", purpose=purpose,
        ))
    try:
        await session.flush()  # 全局唯一约束兜底（勾选后他人抢先登记的竞态）
    except IntegrityError:
        await session.rollback()
        raise AppError(code=ERR_CONFLICT, message="登记冲突（IP 或 MAC 已被占用），请刷新发现结果后重试")
    await record(
        session, user, "ipam:discover:register", target_type="discovery", target_id=str(discovery_id),
        detail={"ips": ips, "count": len(ips), "purpose": purpose, "auto_subnet": auto_subnet},
        ip_address=await get_client_ip(request), user_agent=await get_user_agent(request),
    )
    await session.commit()
    return ok_response(data={"registered": len(ips), "subnet_id": subnet.id})
