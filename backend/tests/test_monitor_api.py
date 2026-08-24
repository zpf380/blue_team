"""监控子系统集成测试：设备 / IPAM / 告警 / 扫描 / 网络发现 / 数据范围 / 权限。"""
import asyncio
import datetime as dt
import ipaddress
import uuid

import pytest
from sqlalchemy import delete, select

from app.models import Alert, Device, DevicePatrol, IPAllocation, IPSubnet, NetworkDiscovery, OperationLog, ScanReport


def _h(token):
    return {"Authorization": f"Bearer {token}"}


async def _login(client, username, password="Bt@123456"):
    resp = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert resp.json()["code"] == 0, resp.json()
    return resp.json()["data"]["access_token"]


def _uniq_ip() -> str:
    return f"10.0.99.{int(uuid.uuid4().hex[:4], 16) % 250 + 1}"


def _scan_ip() -> str:
    """扫描目标必须落在已登记子网内（业务网 10.0.10.0/24）。"""
    return f"10.0.10.{int(uuid.uuid4().hex[:4], 16) % 240 + 1}"


async def _clear_network_residue(session, net: str) -> None:
    """清理目标网段内的测试残留台账，保证发现测试可重复运行（随机网段跨运行可能撞车）。"""
    net = ipaddress.ip_network(net)
    ips = [str(net.network_address + i) for i in range(1, net.num_addresses)]
    await session.execute(delete(IPAllocation).where(IPAllocation.ip_address.in_(ips)))
    await session.execute(delete(Device).where(Device.ip_address.in_(ips)))
    await session.execute(delete(NetworkDiscovery).where(NetworkDiscovery.network == str(net)))
    await session.execute(delete(IPSubnet).where(IPSubnet.network == str(net)))
    await session.commit()


# 模拟 nmap -oX 输出：22(ssh) 与 6379(redis) 开放，443 closed
SAMPLE_XML = """<?xml version="1.0"?>
<nmaprun scanner="nmap" version="7.94">
  <host><status state="up"/>
    <ports>
      <port protocol="tcp" portid="22"><state state="open" reason="syn-ack"/>
        <service name="ssh" product="OpenSSH" version="8.9p1"/></port>
      <port protocol="tcp" portid="6379"><state state="open"/></port>
      <port protocol="tcp" portid="443"><state state="closed"/></port>
    </ports>
  </host>
</nmaprun>
"""


async def _fake_nmap(target, ports, service_detection, scan_options=None):
    return 0, SAMPLE_XML, ""


@pytest.mark.asyncio
async def test_device_crud_and_data_scope(client):
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")
    trainee_t = await _login(client, "trainee01")

    # 学员无监控权限
    resp = await client.get("/api/v1/monitor/devices", headers=_h(trainee_t))
    assert resp.json()["code"] == 40302

    # 数据范围：manager 见全部（>=4），analyst 只见本部门（db-01）
    resp = await client.get("/api/v1/monitor/devices", headers=_h(manager_t), params={"size": 50})
    manager_names = {d["name"] for d in resp.json()["data"]["items"]}
    assert "web-01" in manager_names and "db-01" in manager_names

    resp = await client.get("/api/v1/monitor/devices", headers=_h(analyst_t), params={"size": 50})
    analyst_names = {d["name"] for d in resp.json()["data"]["items"]}
    assert "db-01" in analyst_names
    assert "web-01" not in analyst_names  # web-01 属安全运营部

    # 创建设备（analyst 有 device:manage）
    ip = _uniq_ip()
    resp = await client.post("/api/v1/monitor/devices", headers=_h(analyst_t), json={
        "name": "test-srv", "ip_address": ip, "device_type": "server", "status": "active",
    })
    assert resp.json()["code"] == 0, resp.json()
    dev_id = resp.json()["data"]["id"]

    # 重复 IP → 冲突
    resp = await client.post("/api/v1/monitor/devices", headers=_h(analyst_t), json={"name": "dup", "ip_address": ip})
    assert resp.json()["code"] == 40900

    # ping 更新在线时间
    resp = await client.post(f"/api/v1/monitor/devices/{dev_id}/ping", headers=_h(analyst_t))
    assert resp.json()["code"] == 0
    assert resp.json()["data"]["last_seen_at"]

    # 删除（无告警 → 真删，带删除原因 → 写入审计 detail）
    resp = await client.request("DELETE", f"/api/v1/monitor/devices/{dev_id}", headers=_h(analyst_t), params={"reason": "测试清理"})
    assert resp.json()["data"]["message"] == "设备已删除"


@pytest.mark.asyncio
async def test_device_delete_reason_in_audit(client, test_session):
    """删除设备的删除原因必须落 operation_logs.detail（审计留痕）。"""
    analyst_t = await _login(client, "analyst01")

    ip = _uniq_ip()
    resp = await client.post("/api/v1/monitor/devices", headers=_h(analyst_t), json={
        "name": "del-audit-srv", "ip_address": ip, "device_type": "server", "status": "active",
    })
    assert resp.json()["code"] == 0, resp.json()
    dev_id = resp.json()["data"]["id"]

    resp = await client.request("DELETE", f"/api/v1/monitor/devices/{dev_id}", headers=_h(analyst_t), params={"reason": "设备报废下线"})
    assert resp.json()["data"]["message"] == "设备已删除"

    log = (await test_session.execute(
        select(OperationLog).where(OperationLog.action == "monitor:device:delete")
        .order_by(OperationLog.id.desc()).limit(1)
    )).scalar_one()
    assert log.target_id == str(dev_id)
    assert (log.detail or {}).get("reason") == "设备报废下线"
    assert log.detail.get("archived") is False


@pytest.mark.asyncio
async def test_subnet_allocation_lifecycle(client):
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    # analyst 无 ipam:manage
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(analyst_t), json={"name": "x", "network": "10.0.0.0/24"})
    assert resp.json()["code"] == 40302

    # manager 建子网
    net = f"172.16.{int(uuid.uuid4().hex[:4], 16) % 200 + 1}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "test-net", "network": net, "gateway": net.rsplit(".0/24", 1)[0] + ".1",
    })
    assert resp.json()["code"] == 0, resp.json()
    subnets = (await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))).json()["data"]
    subnet = next(s for s in subnets if s["network"] == net)

    # 非法网段被拒
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "bad", "network": "not-a-cidr"})
    assert resp.json()["code"] == 40001

    # 自动分配首个可用地址
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet["id"], "allocation_type": "static", "purpose": "自动分配测试",
    })
    assert resp.json()["code"] == 0, resp.json()
    first_ip = resp.json()["data"]["ip_address"]

    # 手动分配同样地址 → 冲突
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet["id"], "ip_address": first_ip, "allocation_type": "static",
    })
    assert resp.json()["code"] == 40900

    # 子网外地址被拒
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet["id"], "ip_address": "8.8.8.8", "allocation_type": "static",
    })
    assert resp.json()["code"] == 40001

    # 释放
    alloc_id = (await client.get("/api/v1/monitor/allocations", headers=_h(manager_t), params={"subnet_id": subnet["id"]})).json()["data"]["items"][0]["id"]
    resp = await client.delete(f"/api/v1/monitor/allocations/{alloc_id}", headers=_h(manager_t))
    assert resp.json()["code"] == 0


@pytest.mark.asyncio
async def test_subnet_auto_gateway_and_delete(client, test_session):
    """自动网关推导 / 重复网段拦截 / 有分配拒绝删除 / 释放后删除落审计。"""
    manager_t = await _login(client, "manager01")

    # 只填名称+网段 → 网关自动取第一个可用地址（10.0.30.0/24 → 10.0.30.1）
    net = f"10.255.{int(uuid.uuid4().hex[:4], 16) % 250 + 1}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "auto-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    gw = resp.json()["data"]["gateway"]
    assert gw == net.rsplit(".0/24", 1)[0] + ".1"

    # 同网段重复登记 → 冲突
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "dup-net", "network": net})
    assert resp.json()["code"] == 40900

    subnets = (await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))).json()["data"]
    subnet = next(s for s in subnets if s["network"] == net)

    # 自动分配一个地址 → 有活跃分配时删除子网被拒
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet["id"], "allocation_type": "static", "purpose": "删除拦截测试",
    })
    assert resp.json()["code"] == 0, resp.json()
    alloc_id = resp.json()["data"]["id"]
    resp = await client.request("DELETE", f"/api/v1/monitor/subnets/{subnet['id']}", headers=_h(manager_t), params={"reason": "x"})
    assert resp.json()["code"] == 40900

    # 释放后删除成功，删除原因落审计
    await client.delete(f"/api/v1/monitor/allocations/{alloc_id}", headers=_h(manager_t))
    resp = await client.request("DELETE", f"/api/v1/monitor/subnets/{subnet['id']}", headers=_h(manager_t), params={"reason": "网段规划调整"})
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["message"] == "子网已删除"

    log = (await test_session.execute(
        select(OperationLog).where(OperationLog.action == "ipam:subnet:delete")
        .order_by(OperationLog.id.desc()).limit(1)
    )).scalar_one()
    assert log.target_id == str(subnet["id"])
    assert (log.detail or {}).get("reason") == "网段规划调整"
    assert (log.detail or {}).get("network") == net


@pytest.mark.asyncio
async def test_alert_workflow(client):
    manager_t = await _login(client, "manager01")
    auditor_t = await _login(client, "auditor01")

    # 审计员只读
    resp = await client.get("/api/v1/monitor/alerts", headers=_h(auditor_t))
    assert resp.json()["code"] == 0
    resp = await client.post("/api/v1/monitor/alerts", headers=_h(auditor_t), json={"title": "x", "severity": "high"})
    assert resp.json()["code"] == 40302

    # manager 建告警
    resp = await client.post("/api/v1/monitor/alerts", headers=_h(manager_t), json={
        "title": "测试告警-模拟入侵", "severity": "critical", "alert_type": "intrusion",
        "description": "集成测试用",
    })
    assert resp.json()["code"] == 0, resp.json()
    alert_id = resp.json()["data"]["id"]

    # 确认
    resp = await client.post(f"/api/v1/monitor/alerts/{alert_id}/acknowledge", headers=_h(manager_t))
    assert resp.json()["code"] == 0
    # 重复确认 → 校验失败
    resp = await client.post(f"/api/v1/monitor/alerts/{alert_id}/acknowledge", headers=_h(manager_t))
    assert resp.json()["code"] == 40001
    # 解决
    resp = await client.post(f"/api/v1/monitor/alerts/{alert_id}/resolve", headers=_h(manager_t))
    assert resp.json()["code"] == 0


@pytest.mark.asyncio
async def test_scan_nmap_async_and_review(client, test_session, monkeypatch):
    analyst_t = await _login(client, "analyst01")
    manager_t = await _login(client, "manager01")
    auditor_t = await _login(client, "auditor01")

    # 审计员无扫描权限
    resp = await client.post("/api/v1/monitor/scans", headers=_h(auditor_t), json={"target_ip": "10.0.10.11"})
    assert resp.json()["code"] == 40302

    # 公网目标 → 内网硬校验拒绝（一票否决，优先于授权判定）
    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": "8.8.8.8"})
    assert resp.json()["code"] == 40001
    assert "内网" in resp.json()["message"]

    # 替换子进程边界：扫描编排走真实现，nmap 调用被 mock
    monkeypatch.setattr("app.services.scanner._run_nmap", _fake_nmap)

    # analyst 发起扫描 → 立即返回 pending + report_id（异步）
    # 关联本部门设备 db-01：报告列表已按部门数据范围过滤，无设备报告对 dept 角色不可见
    target = _scan_ip()
    db_dev = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()
    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={
        "target_ip": target, "report_type": "on_demand", "device_id": db_dev.id,
    })
    assert resp.json()["code"] == 0, resp.json()
    data = resp.json()["data"]
    assert "report_id" in data and data["scan_status"] in ("pending", "running")
    report_id = data["report_id"]

    # 轮询至扫描完成（mock 立即返回，通常一轮即达）
    detail = None
    for _ in range(200):
        resp = await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))
        detail = resp.json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "completed", detail
    assert detail["status"] == "pending_review"

    # 真实 XML 解析结果：22(ssh) 与 6379(redis) 开放，443 closed 被过滤
    ports = detail["scan_data"]["open_ports"]
    assert {p["port"] for p in ports} == {22, 6379}
    p22 = next(p for p in ports if p["port"] == 22)
    assert p22["service"] == "ssh" and p22["version"] == "8.9p1"
    # 风险推导：redis critical + ssh low + info 兜底（非标准端口无）
    sevs = {v["port"]: v["severity"] for v in detail["scan_data"]["vulnerabilities"]}
    assert sevs[6379] == "critical" and sevs[22] == "low"
    assert isinstance(detail["risk_score"], int) and 0 < detail["risk_score"] <= 100

    # 报告列表含该报告且带扫描状态
    resp = await client.get("/api/v1/monitor/scans/reports", headers=_h(analyst_t))
    item = next(r for r in resp.json()["data"]["items"] if r["id"] == report_id)
    assert item["scan_status"] == "completed"

    # analyst 无审核权限（仅 manager/admin）
    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/review", headers=_h(analyst_t), params={"approve": "true"})
    assert resp.json()["code"] == 40302

    # manager 审核通过
    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/review", headers=_h(manager_t), params={"approve": "true"})
    assert resp.json()["data"]["status"] == "approved"


@pytest.mark.asyncio
async def test_scan_failure_and_review_guard(client, monkeypatch):
    analyst_t = await _login(client, "analyst01")
    manager_t = await _login(client, "manager01")

    async def _boom(target, ports, svc, scan_options=None):
        raise RuntimeError("boom")

    monkeypatch.setattr("app.services.scanner._run_nmap", _boom)

    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": _scan_ip()})
    assert resp.json()["code"] == 0, resp.json()
    report_id = resp.json()["data"]["report_id"]

    detail = None
    for _ in range(200):
        resp = await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))
        detail = resp.json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "failed"
    assert "boom" in detail["error"]

    # 扫描失败的报告禁止审核
    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/review", headers=_h(manager_t), params={"approve": "true"})
    assert resp.json()["code"] == 40001
    assert "尚未完成" in resp.json()["message"]


@pytest.mark.asyncio
async def test_subnet_overlap_and_nesting_rejected(client):
    """网段重叠/嵌套检测：与已登记网段重叠/包含/被包含一律拒绝，独立网段正常登记。"""
    manager_t = await _login(client, "manager01")

    # seed 办公网为 10.0.0.0/24 → 其子网 10.0.0.0/25 重叠
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "overlap", "network": "10.0.0.0/25"})
    assert resp.json()["code"] == 40900
    assert "重叠" in resp.json()["message"]

    # 包含多个已登记网段的大网段 → 拒绝
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "supernet", "network": "10.0.0.0/16"})
    assert resp.json()["code"] == 40900

    # 独立网段 → 正常登记；随即清理，避免残留网段影响重复运行
    net = f"10.200.{int(uuid.uuid4().hex[:4], 16) % 250 + 1}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "independent", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    independent_id = resp.json()["data"]["id"]
    await client.request("DELETE", f"/api/v1/monitor/subnets/{independent_id}", headers=_h(manager_t), params={"reason": "test cleanup"})


@pytest.mark.asyncio
async def test_ipam_department_data_scope(client):
    """IPAM 部门数据范围：analyst 仅见本部门（攻防实验室）子网与分配，manager 全量。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    # manager 见全部 seed 子网
    resp = await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))
    manager_nets = {s["network"] for s in resp.json()["data"]}
    assert {"10.0.0.0/24", "10.0.10.0/24", "10.0.20.0/24"} <= manager_nets

    # analyst 只见业务网（攻防实验室），办公网/服务器网不可见
    resp = await client.get("/api/v1/monitor/subnets", headers=_h(analyst_t))
    analyst_nets = {s["network"] for s in resp.json()["data"]}
    assert "10.0.10.0/24" in analyst_nets
    assert "10.0.0.0/24" not in analyst_nets
    assert "10.0.20.0/24" not in analyst_nets

    # manager 在办公网（安全运营部）手动分配一个随机地址
    office = next(s for s in (await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))).json()["data"] if s["network"] == "10.0.0.0/24")
    office_ip = f"10.0.0.{int(uuid.uuid4().hex[:4], 16) % 200 + 50}"
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": office["id"], "ip_address": office_ip, "allocation_type": "static", "purpose": "数据范围测试",
    })
    assert resp.json()["code"] == 0, resp.json()

    # analyst 看不到办公网分配，但能看到业务网 seed 分配
    resp = await client.get("/api/v1/monitor/allocations", headers=_h(analyst_t), params={"size": 50})
    analyst_ips = {a["ip_address"] for a in resp.json()["data"]["items"]}
    assert office_ip not in analyst_ips
    assert "10.0.10.11" in analyst_ips

    # manager 能看到办公网分配
    resp = await client.get("/api/v1/monitor/allocations", headers=_h(manager_t), params={"size": 50})
    assert office_ip in {a["ip_address"] for a in resp.json()["data"]["items"]}

    # 清理本次创建的数据范围测试分配：若不删，每次运行都在 seed 办公网累积一条 10.0.0.x 残留，
    # 随机 office_ip 撞旧残留 → 409 → flaky（conftest 的 purpose='数据范围测试' 清理兜底）
    office_alloc = next(a for a in resp.json()["data"]["items"] if a["ip_address"] == office_ip)
    await client.delete(f"/api/v1/monitor/allocations/{office_alloc['id']}", headers=_h(manager_t))


@pytest.mark.asyncio
async def test_expired_lease_recycled(client):
    """过期 DHCP 租约自动回收：过期租约从列表消失、地址可复用；static/reserved 不受影响。"""
    manager_t = await _login(client, "manager01")
    # 双段随机（第二段 210-239 避开种子 0/10/20 与 200-209 既有测试残留区），
    # 消除单段随机跨运行撞残留子网的 flaky（该测试末尾不删子网，会累积残留）
    b2 = 210 + int(uuid.uuid4().hex[:2], 16) % 30
    c = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.{b2}.{c}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "lease-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    past = (dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=1)).isoformat()
    future = (dt.datetime.now(dt.timezone.utc) + dt.timedelta(days=1)).isoformat()

    # 过期 DHCP 租约（自动分配，应为子网第一个可用地址）
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "dhcp", "purpose": "过期租约", "expires_at": past,
    })
    assert resp.json()["code"] == 0, resp.json()
    expired_ip = resp.json()["data"]["ip_address"]

    # 列表查询触发惰性回收 → 过期租约消失
    resp = await client.get("/api/v1/monitor/allocations", headers=_h(manager_t), params={"subnet_id": subnet_id})
    assert expired_ip not in {a["ip_address"] for a in resp.json()["data"]["items"]}

    # 被回收地址可立即重新分配（物理删除而非软删）
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "ip_address": expired_ip, "allocation_type": "static", "purpose": "复用测试",
    })
    assert resp.json()["code"] == 0, resp.json()

    # 未过期 DHCP 租约保留
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "dhcp", "purpose": "有效租约", "expires_at": future,
    })
    assert resp.json()["code"] == 0, resp.json()

    # 过期 static 不被回收
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "static", "purpose": "过期静态", "expires_at": past,
    })
    assert resp.json()["code"] == 0, resp.json()
    stale_static_ip = resp.json()["data"]["ip_address"]

    resp = await client.get("/api/v1/monitor/allocations", headers=_h(manager_t), params={"subnet_id": subnet_id})
    active_ips = {a["ip_address"] for a in resp.json()["data"]["items"]}
    assert stale_static_ip in active_ips
    assert expired_ip in active_ips  # 复用后的 static 记录仍在


@pytest.mark.asyncio
async def test_subnet_reserved_ranges(client):
    """保留地址段：自动分配跳过保留段，手动指定不受限，非法保留段被拒。"""
    manager_t = await _login(client, "manager01")
    x = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.202.{x}.0/24"
    reserved_input = f"10.202.{x}.100/28"  # 主机位非零 → 归一化为网络地址 .96/28
    reserved = [f"10.202.{x}.96/28"]  # 覆盖 .96-.111，自动分配不触及

    # 非法保留段（不在子网内）→ 校验失败
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "res-bad", "network": f"10.203.{x}.0/24", "reserved_ranges": ["8.8.8.0/24"],
    })
    assert resp.json()["code"] == 40001
    assert "不在子网" in resp.json()["message"]

    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "res-net", "network": net, "reserved_ranges": [reserved_input],
    })
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    # 自动分配跳过保留段：gateway .1，应拿 .2
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "static", "purpose": "auto",
    })
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["ip_address"] == f"10.202.{x}.2"

    # 手动分配保留段内地址 → 允许
    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "ip_address": f"10.202.{x}.100", "allocation_type": "reserved", "purpose": "保留服务器",
    })
    assert resp.json()["code"] == 0, resp.json()

    # list_subnets 带保留段信息
    resp = await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))
    s = next(s for s in resp.json()["data"] if s["id"] == subnet_id)
    assert s["reserved_ranges"] == reserved


@pytest.mark.asyncio
async def test_subnet_edit_and_update_audit(client, test_session):
    """子网编辑：改名/网关/VLAN/保留段生效，非法网关与跨段保留段被拒，审计留痕。"""
    manager_t = await _login(client, "manager01")
    x = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.203.{x}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "edit-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    # 非法网关（不在网段内）→ 拒绝
    resp = await client.put(f"/api/v1/monitor/subnets/{subnet_id}", headers=_h(manager_t), json={"gateway": "8.8.8.8"})
    assert resp.json()["code"] == 40001

    # 跨段保留段 → 拒绝
    resp = await client.put(f"/api/v1/monitor/subnets/{subnet_id}", headers=_h(manager_t), json={"reserved_ranges": [f"10.204.{x}.0/24"]})
    assert resp.json()["code"] == 40001

    # 正常编辑（保留段 .100/28 归一化为 .96/28）
    resp = await client.put(f"/api/v1/monitor/subnets/{subnet_id}", headers=_h(manager_t), json={
        "name": "edit-net-renamed", "gateway": f"10.203.{x}.5", "vlan_id": 99,
        "reserved_ranges": [f"10.203.{x}.100/28"],
    })
    assert resp.json()["code"] == 0, resp.json()

    resp = await client.get("/api/v1/monitor/subnets", headers=_h(manager_t))
    s = next(s for s in resp.json()["data"] if s["id"] == subnet_id)
    assert s["name"] == "edit-net-renamed" and s["gateway"] == f"10.203.{x}.5"
    assert s["vlan_id"] == 99 and s["reserved_ranges"] == [f"10.203.{x}.96/28"]

    log = (await test_session.execute(
        select(OperationLog).where(OperationLog.action == "ipam:subnet:update")
        .order_by(OperationLog.id.desc()).limit(1)
    )).scalar_one()
    assert log.target_id == str(subnet_id)
    assert (log.detail or {}).get("changes", {}).get("name") == "edit-net-renamed"


@pytest.mark.asyncio
async def test_allocation_edit_and_history(client, test_session):
    """分配记录编辑 + 按 IP 查审计轨迹。"""
    manager_t = await _login(client, "manager01")
    x = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.204.{x}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "hist-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "static", "purpose": "原始用途",
    })
    assert resp.json()["code"] == 0, resp.json()
    alloc_id = resp.json()["data"]["id"]
    ip = resp.json()["data"]["ip_address"]

    # 编辑用途与类型
    resp = await client.put(f"/api/v1/monitor/allocations/{alloc_id}", headers=_h(manager_t), json={
        "purpose": "新用途", "allocation_type": "reserved",
    })
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["changes"]["purpose"] == "新用途"

    # 无变更
    resp = await client.put(f"/api/v1/monitor/allocations/{alloc_id}", headers=_h(manager_t), json={"purpose": "新用途"})
    assert resp.json()["data"]["message"] == "无变更"

    # 按 IP 查历史：应有 create + update 两条
    resp = await client.get("/api/v1/monitor/allocations/history", headers=_h(manager_t), params={"ip": ip})
    assert resp.json()["code"] == 0, resp.json()
    actions = [h["action"] for h in resp.json()["data"]]
    assert "ipam:alloc:create" in actions and "ipam:alloc:update" in actions

    log = (await test_session.execute(
        select(OperationLog).where(OperationLog.action == "ipam:alloc:update")
        .order_by(OperationLog.id.desc()).limit(1)
    )).scalar_one()
    assert (log.detail or {}).get("ip") == ip


@pytest.mark.asyncio
async def test_subnet_usage_heatmap_and_scope(client):
    """子网使用明细接口：返回分配明细；越权访问他部门子网被拒。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")
    x = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.205.{x}.0/24"

    # 属安全运营部（analyst 的攻防实验室不可见）
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "usage-net", "network": net, "department_id": 1, "reserved_ranges": [f"10.205.{x}.200/30"],
    })
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
        "subnet_id": subnet_id, "allocation_type": "static", "purpose": "热图测试",
    })
    assert resp.json()["code"] == 0, resp.json()
    used_ip = resp.json()["data"]["ip_address"]

    resp = await client.get(f"/api/v1/monitor/subnets/{subnet_id}/usage", headers=_h(manager_t))
    data = resp.json()["data"]
    assert data["used"] == 1 and data["capacity"] == 256
    assert data["reserved_ranges"] == [f"10.205.{x}.200/30"]
    assert any(a["ip"] == used_ip for a in data["allocations"])

    # analyst 越权查看 → 数据范围拒绝（40301）
    resp = await client.get(f"/api/v1/monitor/subnets/{subnet_id}/usage", headers=_h(analyst_t))
    assert resp.json()["code"] == 40301


@pytest.mark.asyncio
async def test_discovery_lifecycle_and_register(client, test_session, monkeypatch):
    """网络发现：手动网段→带 MAC 扫描→幽灵分组→登记为终端设备+DHCP分配+自动建子网。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    # 随机 /28 网段，避免重复跑测试撞车；不预登记子网（走手动网段路径）
    a = 206 + int(uuid.uuid4().hex[:2], 16) % 3
    b = int(uuid.uuid4().hex[:4], 16) % 250
    net = f"10.{a}.{b}.0/28"
    await _clear_network_residue(test_session, net)
    base = ipaddress.ip_network(net).network_address

    # 预置台账设备：.4 在线已登记、.5 已登记但扫描不报（将推导为 offline）
    for off in (4, 5):
        test_session.add(Device(
            name=f"预置-{off}", ip_address=str(base + off), mac_address=f"AA:BB:CC:DD:EE:0{off}", status="active",
        ))
    await test_session.commit()

    # 模拟 nmap -sn 输出：.2(带随机MAC)/.3(无MAC)/.4(带MAC) 在线，.5 不报
    ghost_mac = f"02:00:00:{uuid.uuid4().hex[:2]}:{uuid.uuid4().hex[:2]}:{uuid.uuid4().hex[:2]}"
    async def _fake_discovery(network):
        h2 = (f'<host><status state="up"/><address addr="{str(base + 2)}" addrtype="ipv4"/>'
              f'<address addr="{ghost_mac}" addrtype="mac" vendor="VMware, Inc."/></host>')
        h3 = f'<host><status state="up"/><address addr="{str(base + 3)}" addrtype="ipv4"/></host>'
        h4 = (f'<host><status state="up"/><address addr="{str(base + 4)}" addrtype="ipv4"/>'
              f'<address addr="AA:BB:CC:DD:EE:04" addrtype="mac"/></host>')
        return 0, f'<?xml version="1.0"?><nmaprun scanner="nmap" version="7.94">{h2}{h3}{h4}</nmaprun>'
    monkeypatch.setattr("app.services.scanner._run_host_discovery", _fake_discovery)

    # analyst 无 ipam:manage → 40302（权限校验先于数据范围）
    resp = await client.post("/api/v1/monitor/discover", headers=_h(analyst_t), json={"network": net})
    assert resp.json()["code"] == 40302

    # 手动网段（未登记子网）须先登记扫描授权；非内网/未授权网段发起都会被拒
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "lifecycle-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()

    # manager 手动网段发起 → 立即返回 pending + discovery_id（异步后台执行）
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": net})
    assert resp.json()["code"] == 0, resp.json()
    did = resp.json()["data"]["discovery_id"]
    assert resp.json()["data"]["scan_status"] == "pending"

    # 轮询至完成（mock 立即返回）
    detail = None
    for _ in range(200):
        resp = await client.get(f"/api/v1/monitor/discover/{did}", headers=_h(manager_t))
        detail = resp.json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "completed", detail
    assert detail["netmask"] == "255.255.255.240"  # /28 掩码
    # hosts 元数据：.2 带 MAC+vendor、.3 无 MAC、.4 带 MAC
    host_map = {h["ip"]: h for h in detail["hosts"]}
    assert host_map[str(base + 2)]["mac"] == ghost_mac
    assert host_map[str(base + 2)]["vendor"] == "VMware, Inc."
    assert host_map[str(base + 3)]["mac"] is None
    assert detail["online_ips"] == [str(base + i) for i in (2, 3, 4)]
    assert detail["unregistered_ips"] == [str(base + i) for i in (2, 3)]  # 幽灵设备
    assert detail["registered_ips"] == [str(base + 4)]
    assert detail["offline_ips"] == [str(base + 5)]

    # 勾选登记幽灵设备 .2/.3 → 创建终端设备 + DHCP 分配 + 自动建子网
    resp = await client.post(f"/api/v1/monitor/discover/{did}/register", headers=_h(manager_t),
                             json={"ips": [str(base + 2), str(base + 3)], "purpose": "发现登记"})
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["registered"] == 2
    auto_subnet_id = resp.json()["data"]["subnet_id"]
    # 自动创建了对应子网（掩码固化进 IPAM 台账）
    sub = await test_session.get(IPSubnet, auto_subnet_id)
    assert sub is not None and str(sub.network) == net
    assert "终端网段" in sub.name
    # 审计落库，含 auto_subnet 标记
    log = (await test_session.execute(
        select(OperationLog).where(OperationLog.action == "ipam:discover:register")
        .order_by(OperationLog.id.desc()).limit(1)
    )).scalar_one()
    assert log.target_id == str(did)
    assert (log.detail or {}).get("count") == 2
    assert (log.detail or {}).get("auto_subnet") is True

    # 登记生成了设备记录（IP + MAC + 厂商）
    dev2 = (await test_session.execute(select(Device).where(Device.ip_address == str(base + 2)))).scalar_one()
    assert str(dev2.mac_address).lower() == ghost_mac.lower()
    assert dev2.manufacturer == "VMware, Inc."
    assert dev2.name == f"终端-{str(base + 2)}"
    dev3 = (await test_session.execute(select(Device).where(Device.ip_address == str(base + 3)))).scalar_one()
    assert dev3.mac_address is None
    # DHCP 分配关联设备与自动子网
    alloc2 = (await test_session.execute(select(IPAllocation).where(IPAllocation.ip_address == str(base + 2)))).scalar_one()
    assert alloc2.device_id == dev2.id and alloc2.subnet_id == auto_subnet_id
    assert alloc2.allocation_type == "dhcp" and alloc2.purpose == "发现登记"

    # 重复登记同 IP → 40900（已登记为设备）
    resp = await client.post(f"/api/v1/monitor/discover/{did}/register", headers=_h(manager_t), json={"ips": [str(base + 2)]})
    assert resp.json()["code"] == 40900
    # 非法 IP（不在网段）→ 40001
    resp = await client.post(f"/api/v1/monitor/discover/{did}/register", headers=_h(manager_t), json={"ips": ["10.99.99.99"]})
    assert resp.json()["code"] == 40001

    # 数据范围：analyst 看不到该无部门网段的发现记录
    resp = await client.get("/api/v1/monitor/discover", headers=_h(manager_t))
    assert any(item["id"] == did for item in resp.json()["data"]["items"])
    resp = await client.get("/api/v1/monitor/discover", headers=_h(analyst_t))
    assert not any(item["id"] == did for item in resp.json()["data"]["items"])


@pytest.mark.asyncio
async def test_discovery_guards(client, test_session, monkeypatch):
    """发现守卫：非法 CIDR / 超大网段 / 未完成登记 / 越权 / MAC 冲突。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    # 非法 CIDR → 40001
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": "not-a-network"})
    assert resp.json()["code"] == 40001
    assert "格式不正确" in resp.json()["message"]

    # 公网网段 → 内网硬校验拒绝（一票否决，先于授权判定）
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": "8.8.8.0/24"})
    assert resp.json()["code"] == 40001
    assert "内网" in resp.json()["message"]

    # 超大网段（/16 65536 地址 > 1024）→ 40001（内网 /16，资源保护优先于授权判定）
    big = f"10.{int(uuid.uuid4().hex[:2], 16) % 200}.0.0/16"
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": big})
    assert resp.json()["code"] == 40001
    assert "网段过大" in resp.json()["message"]

    # 内网但未授权网段 → 拒绝（不在登记子网亦不在授权名单）
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": "10.244.0.0/28"})
    assert resp.json()["code"] == 40001
    assert "授权" in resp.json()["message"]

    # 卡住的主机发现，让任务停在 pending/running
    async def _slow_discovery(network):
        await asyncio.sleep(1.0)
        return 0, '<?xml version="1.0"?><nmaprun/>'
    monkeypatch.setattr("app.services.scanner._run_host_discovery", _slow_discovery)

    a = 209 + int(uuid.uuid4().hex[:2], 16) % 1
    b = int(uuid.uuid4().hex[:4], 16) % 250
    net = f"10.{a}.{b}.0/28"
    await _clear_network_residue(test_session, net)
    # 手动网段须先登记扫描授权
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "guards-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    # 发起后立即登记（任务尚未完成）→ 40001
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": net})
    assert resp.json()["code"] == 0, resp.json()
    did = resp.json()["data"]["discovery_id"]
    resp = await client.post(f"/api/v1/monitor/discover/{did}/register", headers=_h(manager_t), json={"ips": ["10.0.0.1"]})
    assert resp.json()["code"] == 40001
    assert "尚未完成" in resp.json()["message"]

    # analyst 无 ipam:manage，权限校验（40302）先于数据范围；manager 对任何子网均可发起
    net2 = f"10.{a}.{int(uuid.uuid4().hex[:4], 16) % 250}.0/28"
    await _clear_network_residue(test_session, net2)
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "disc-dept1", "network": net2, "department_id": 1,
    })
    assert resp.json()["code"] == 0, resp.json()
    dept1_id = resp.json()["data"]["id"]
    resp = await client.post("/api/v1/monitor/discover", headers=_h(analyst_t), json={"subnet_id": dept1_id})
    assert resp.json()["code"] == 40302
    # manager 对部门1 子网发起（关联子网路径）→ 权限+数据范围均通过
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"subnet_id": dept1_id})
    assert resp.json()["code"] == 0, resp.json()

    # MAC 冲突：预置设备占用某 MAC，扫描幽灵携带同 MAC → 登记 40900（MAC 随机，避免跨运行撞车）
    base2 = ipaddress.ip_network(net2).network_address
    ghost_mac = f"02:00:00:{uuid.uuid4().hex[:2]}:{uuid.uuid4().hex[:2]}:{uuid.uuid4().hex[:2]}"
    test_session.add(Device(
        name="mac-owner", ip_address=str(base2 + 9), mac_address=ghost_mac, status="active",
    ))
    await test_session.commit()

    async def _fake_with_mac(network):
        h2 = f'<host><status state="up"/><address addr="{str(base2 + 2)}" addrtype="ipv4"/></host>'
        h3 = (f'<host><status state="up"/><address addr="{str(base2 + 3)}" addrtype="ipv4"/>'
              f'<address addr="{ghost_mac}" addrtype="mac" vendor="Test Vendor"/></host>')
        return 0, f'<?xml version="1.0"?><nmaprun scanner="nmap" version="7.94">{h2}{h3}</nmaprun>'
    monkeypatch.setattr("app.services.scanner._run_host_discovery", _fake_with_mac)

    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"subnet_id": dept1_id})
    assert resp.json()["code"] == 0, resp.json()
    did2 = resp.json()["data"]["discovery_id"]
    for _ in range(200):
        r = await client.get(f"/api/v1/monitor/discover/{did2}", headers=_h(manager_t))
        if r.json()["data"]["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    resp = await client.post(f"/api/v1/monitor/discover/{did2}/register", headers=_h(manager_t),
                             json={"ips": [str(base2 + 3)]})
    assert resp.json()["code"] == 40900
    assert "MAC" in resp.json()["message"]

    # 等慢任务跑完，避免后台任务跨测试泄漏
    for _ in range(100):
        resp = await client.get(f"/api/v1/monitor/discover/{did}", headers=_h(manager_t))
        if resp.json()["data"]["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)


@pytest.mark.asyncio
async def test_scan_authorization_lifecycle(client, test_session, monkeypatch):
    """扫描授权名单：内网硬限制 / 登记 / 未授权拒 / 授权放行 / 吊销拒 / 过期拒 / 越权拒。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    # 公网网段不可授权
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "pub", "network": "8.8.8.0/24"})
    assert resp.json()["code"] == 40001
    assert "内网" in resp.json()["message"]

    # 非法 CIDR → 拒
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "bad", "network": "not-a-cidr"})
    assert resp.json()["code"] == 40001

    # analyst 无授权管理权限（manager/admin）
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(analyst_t), json={"name": "x", "network": "10.0.0.0/8"})
    assert resp.json()["code"] == 40302

    # mock 主机发现子进程边界，避免授权放行后真实跑 nmap
    async def _fake_disc(network):
        return 0, '<?xml version="1.0"?><nmaprun/>'
    monkeypatch.setattr("app.services.scanner._run_host_discovery", _fake_disc)

    # 登记内网授权网段（244 前缀避开种子子网与其它测试残留）
    b = int(uuid.uuid4().hex[:4], 16) % 249 + 1
    net = f"10.244.{b}.0/28"
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "auth-test", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    auth_id = resp.json()["data"]["id"]

    # 列表可见（analyst 有 monitor:scan 可读）
    lst = await client.get("/api/v1/monitor/scan-auth", headers=_h(analyst_t))
    auth_map = {x["id"]: x for x in lst.json()["data"]}
    assert auth_id in auth_map and auth_map[auth_id]["status"] == "active"
    assert auth_map[auth_id]["approved_by_name"]  # 批准人回显

    # 未授权内网网段发起发现 → 拒绝
    net2 = f"10.244.{b + 1 if b < 249 else 1}.0/28"
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": net2})
    assert resp.json()["code"] == 40001
    assert "授权" in resp.json()["message"]

    # 已授权网段 → 放行
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": net})
    assert resp.json()["code"] == 0, resp.json()

    # 吊销后 → 拒绝；重复吊销 → 409
    resp = await client.post(f"/api/v1/monitor/scan-auth/{auth_id}/revoke", headers=_h(manager_t))
    assert resp.json()["code"] == 0
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": net})
    assert resp.json()["code"] == 40001
    assert "授权" in resp.json()["message"]
    resp = await client.post(f"/api/v1/monitor/scan-auth/{auth_id}/revoke", headers=_h(manager_t))
    assert resp.json()["code"] == 40900

    # 过期授权 → 拒绝（end_date 已过）
    past_net = f"10.244.{b + 2 if b < 248 else 1}.0/28"
    past_end = (dt.datetime.now(dt.timezone.utc) - dt.timedelta(minutes=5)).isoformat()
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={"name": "expired", "network": past_net, "end_date": past_end})
    assert resp.json()["code"] == 0, resp.json()
    resp = await client.post("/api/v1/monitor/discover", headers=_h(manager_t), json={"network": past_net})
    assert resp.json()["code"] == 40001

    # 生效时间晚于到期时间 → 拒
    resp = await client.post("/api/v1/monitor/scan-auth", headers=_h(manager_t), json={
        "name": "bad-range", "network": "10.244.250.0/28",
        "start_date": (dt.datetime.now(dt.timezone.utc) + dt.timedelta(days=2)).isoformat(),
        "end_date": (dt.datetime.now(dt.timezone.utc) + dt.timedelta(days=1)).isoformat(),
    })
    assert resp.json()["code"] == 40001

    # 审计留痕：登记与吊销均落库
    logs = (await test_session.execute(
        select(OperationLog).where(OperationLog.action.in_(["ipam:scan_auth:create", "ipam:scan_auth:revoke"]))
    )).scalars().all()
    assert any(l.action == "ipam:scan_auth:create" for l in logs)
    assert any(l.action == "ipam:scan_auth:revoke" for l in logs)


@pytest.mark.asyncio
async def test_scan_concurrency_limited(client, test_session, monkeypatch):
    """扫描并发信号量：同时在跑的 nmap 不超过 SCAN_MAX_CONCURRENT（8 并发被压到上限）。"""
    from app.core.config import settings
    from app.models import ScanReport
    from app.services import scanner as scanner_mod

    active = 0
    peak = 0
    lock = asyncio.Lock()

    async def _slow_nmap(target, ports, svc, scan_options=None):
        nonlocal active, peak
        async with lock:
            active += 1
            peak = max(peak, active)
        await asyncio.sleep(0.05)
        async with lock:
            active -= 1
        return 0, SAMPLE_XML, ""

    monkeypatch.setattr("app.services.scanner._run_nmap", _slow_nmap)

    ids = []
    for i in range(8):
        r = ScanReport(target_ip=f"10.0.10.1{i}", report_type="on_demand", scan_status="pending", generated_by=1)
        test_session.add(r)
        await test_session.flush()
        ids.append(r.id)
    await test_session.commit()

    await asyncio.gather(*[scanner_mod.execute_scan(i, "10.0.10.11", 100) for i in ids])

    assert peak <= settings.SCAN_MAX_CONCURRENT
    assert peak == settings.SCAN_MAX_CONCURRENT  # 8 并发确实被压到上限，证明信号量生效


@pytest.mark.asyncio
async def test_client_error_report(client, test_session):
    """前端错误上报：匿名可上报；Bearer 认证后记录 user_id；字段限长截断。"""
    from app.models import ClientErrorReport

    # 匿名上报（无任何凭证）
    resp = await client.post("/api/v1/monitor/client-errors", json={
        "message": "[error] boom", "url": "http://x/page", "stack": "at f (a.js:1:2)",
    })
    assert resp.json()["code"] == 0

    # 认证上报（Bearer）→ 记录 user_id
    manager_t = await _login(client, "manager01")
    long_stack = "x" * 15000  # 合法（前端已截断，后端 Pydantic max_length=20000 兜底）
    resp = await client.post("/api/v1/monitor/client-errors",
                             json={"message": "TypeError: x is undefined", "stack": long_stack},
                             headers={"Authorization": f"Bearer {manager_t}"})
    assert resp.json()["code"] == 0

    # 超长 message / stack 被 Pydantic 拒收（422 + 40001），防错误风暴打爆库
    resp = await client.post("/api/v1/monitor/client-errors",
                             json={"message": "t" * 1001, "stack": "z"},
                             headers={"Authorization": f"Bearer {manager_t}"})
    assert resp.status_code == 422
    assert resp.json()["code"] == 40001
    resp = await client.post("/api/v1/monitor/client-errors",
                             json={"message": "too long", "stack": "z" * 30000},
                             headers={"Authorization": f"Bearer {manager_t}"})
    assert resp.status_code == 422

    rows = (await test_session.execute(select(ClientErrorReport).order_by(ClientErrorReport.id.desc()).limit(2))).scalars().all()
    assert rows[0].user_id  # 认证后的上报带 user_id
    assert rows[1].user_id is None  # 匿名
    assert rows[0].message == "TypeError: x is undefined"
    assert len(rows[0].stack or "") == 15000  # 合法 stack 原样入库


# ---------- 阶段5：告警通知 / 租约回收 / Excel 导入导出 / 文件白名单 ----------

_HIGH_RISK_XML = """<?xml version="1.0"?>
<nmaprun scanner="nmap" version="7.94">
  <host><status state="up"/>
    <ports>
      <port protocol="tcp" portid="6379"><state state="open"/></port>
      <port protocol="tcp" portid="3306"><state state="open"/></port>
      <port protocol="tcp" portid="3389"><state state="open"/></port>
      <port protocol="tcp" portid="23"><state state="open"/></port>
      <port protocol="tcp" portid="445"><state state="open"/></port>
      <port protocol="tcp" portid="1433"><state state="open"/></port>
      <port protocol="tcp" portid="5900"><state state="open"/></port>
      <port protocol="tcp" portid="9200"><state state="open"/></port>
    </ports>
  </host>
</nmaprun>
"""


@pytest.mark.asyncio
async def test_scan_high_risk_auto_alert(client, test_session, monkeypatch):
    """扫描评分达阈值 → 自动产生告警（含 critical 漏洞 → intrusion/critical）并触发外部通知。"""
    from app.models import Alert, ScanReport
    from app.services import scanner as scanner_mod

    notified = []

    async def _high_nmap(target, ports, svc, scan_options=None):
        return 0, _HIGH_RISK_XML, ""

    async def _fake_notify(alert_id, title, content, severity):
        notified.append((alert_id, title, severity))

    monkeypatch.setattr("app.services.scanner._run_nmap", _high_nmap)
    monkeypatch.setattr("app.services.scanner.notify_alert_task", _fake_notify)

    # 双段随机目标：10.99.{x}.11 仅 200 个取值，跨运行残留的扫描告警会撞随机目标 → dedup 短路 → flaky。
    # 双段全随机（约 6.4 万取值）+ conftest 会话启动清理扫描告警，消除跨运行/测试间碰撞
    target = f"10.{int(uuid.uuid4().hex[:4], 16) % 254 + 1}.{int(uuid.uuid4().hex[:4], 16) % 254 + 1}.11"
    r = ScanReport(target_ip=target, report_type="on_demand", scan_status="pending", generated_by=1)
    test_session.add(r)
    await test_session.commit()

    await scanner_mod.execute_scan(r.id, target, 100)

    rows = (await test_session.execute(select(Alert).where(Alert.title.like(f"%{target}%")))).scalars().all()
    assert len(rows) == 1
    a = rows[0]
    assert a.alert_type == "intrusion" and a.severity == "critical"
    assert str(a.title) == f"扫描发现高风险：{target}（评分 100）"
    assert "Redis 未授权访问" in (a.description or "")
    # 外部通知任务已触发并携带告警 id 与级别
    assert notified and notified[0][0] == a.id and notified[0][2] == "critical"


@pytest.mark.asyncio
async def test_alert_notification_stamps_notified_at(client, test_session, monkeypatch):
    """手动创建告警 → 后台通知任务发送成功后回写 notified_at。"""
    from app.models import Alert

    async def _fake_send(title, content, severity):
        return True

    monkeypatch.setattr("app.services.notify.send_alert_notification", _fake_send)

    manager_t = await _login(client, "manager01")
    resp = await client.post("/api/v1/monitor/alerts", headers=_h(manager_t), json={
        "title": "测试通知闭环", "severity": "high", "alert_type": "abnormal", "description": "desc",
    })
    assert resp.json()["code"] == 0, resp.json()
    alert_id = resp.json()["data"]["id"]

    # 后台任务异步回写，轮询等待（expire_all 冲掉 test_session 身份映射，读到后台提交的新值）
    a = None
    for _ in range(100):
        test_session.expire_all()  # expire_all 是同步方法
        a = await test_session.get(Alert, alert_id)
        if a and a.notified_at:
            break
        await asyncio.sleep(0.05)
    assert a is not None and a.notified_at is not None

    # 列表响应带 notified_at，供前端展示"已通知"
    resp = await client.get("/api/v1/monitor/alerts", headers=_h(manager_t), params={"page": 1, "size": 50})
    item = next(i for i in resp.json()["data"]["items"] if i["id"] == alert_id)
    assert item["notified_at"]


@pytest.mark.asyncio
async def test_recycle_expired_leases_manual(client, test_session):
    """手动回收端点：过期 DHCP 删除、未过期 DHCP 与过期 static 保留，审计留痕。"""
    manager_t = await _login(client, "manager01")
    b2 = 210 + int(uuid.uuid4().hex[:2], 16) % 30
    c = int(uuid.uuid4().hex[:4], 16) % 250 + 1
    net = f"10.{b2}.{c}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "recycle-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]

    # 直插 3 条分配（不用 API：每次 create_allocation 都会触发惰性回收，会把过期租约先清掉）
    now = dt.datetime.now(dt.timezone.utc)
    expired_ip, valid_ip, stale_ip = f"10.{b2}.{c}.1", f"10.{b2}.{c}.2", f"10.{b2}.{c}.3"
    test_session.add_all([
        IPAllocation(subnet_id=subnet_id, ip_address=expired_ip, allocation_type="dhcp", purpose="过期",
                     expires_at=now - dt.timedelta(days=1), is_active=True),
        IPAllocation(subnet_id=subnet_id, ip_address=valid_ip, allocation_type="dhcp", purpose="有效",
                     expires_at=now + dt.timedelta(days=1), is_active=True),
        IPAllocation(subnet_id=subnet_id, ip_address=stale_ip, allocation_type="static", purpose="过期静态",
                     expires_at=now - dt.timedelta(days=1), is_active=True),
    ])
    await test_session.commit()

    resp = await client.post("/api/v1/monitor/allocations/recycle", headers=_h(manager_t))
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["recycled"] == 1

    active = {a["ip_address"] for a in (await client.get(
        "/api/v1/monitor/allocations", headers=_h(manager_t), params={"subnet_id": subnet_id})).json()["data"]["items"]}
    assert expired_ip not in active
    assert valid_ip in active and stale_ip in active

    # 审计留痕（audit 日志仅 admin/auditor 可查）
    auditor_t = await _login(client, "auditor01")
    resp = await client.get("/api/v1/audit/logs", headers=_h(auditor_t), params={"action": "ipam:lease:recycle", "size": 10})
    assert resp.json()["code"] == 0, resp.json()
    assert any(l["action"] == "ipam:lease:recycle" for l in resp.json()["data"]["items"])


@pytest.mark.asyncio
async def test_devices_export_xlsx(client):
    """设备清单导出为有效 XLSX（ZIP 魔数），列与数据完整。"""
    manager_t = await _login(client, "manager01")
    resp = await client.get("/api/v1/monitor/devices/export", headers=_h(manager_t))
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    data = resp.content
    assert data[:2] == b"PK"  # xlsx 是 ZIP 容器

    from openpyxl import load_workbook
    import io as _io

    wb = load_workbook(_io.BytesIO(data), read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows())]
    assert header[0] == "名称" and "IP地址" in header
    names = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert "web-01" in names  # seed 设备


@pytest.mark.asyncio
async def test_devices_import_csv(client, test_session):
    """CSV 导入：合法行创建设备、重复 IP 报错、非法状态报错；导入结果与审计落库。"""
    from app.models import Device

    manager_t = await _login(client, "manager01")
    # 10.7 段：避开 seed(10.0.x)、发现登记设备残留(10.2xx)
    ip_a = f"10.7.{int(uuid.uuid4().hex[:4], 16) % 250 + 1}.5"
    ip_b = f"10.7.{int(uuid.uuid4().hex[:4], 16) % 250 + 1}.6"
    csv_text = (
        "name,ip_address,mac_address,device_type,manufacturer,model,location,department,status\n"
        f"导入终端A,{ip_a},AA:BB:CC:DD:EE:01,workstation,,,机房B,,active\n"
        f"web-01,10.0.10.11,,,,,,,active\n"   # seed 已存在 IP → 冲突
        f"导入终端C,{ip_b},,,,,,,bad_status\n"  # 非法状态
    )
    try:
        resp = await client.post("/api/v1/monitor/devices/import", headers=_h(manager_t),
                                 files={"file": ("devices.csv", csv_text.encode("utf-8-sig"), "text/csv")})
        assert resp.json()["code"] == 0, resp.json()
        data = resp.json()["data"]
        assert data["created"] == 1
        assert any("已存在" in e["error"] for e in data["failed"])
        assert any("非法" in e["error"] for e in data["failed"])

        rows = (await test_session.execute(select(Device).where(Device.ip_address == ip_a))).scalars().all()
        # MACADDR 列由 PG 规范化存储为小写
        assert len(rows) == 1 and rows[0].name == "导入终端A"
        assert (rows[0].mac_address or "").lower() == "aa:bb:cc:dd:ee:01"
    finally:
        # 自清理：导入测试在 DB 中留下 10.7.x 设备，跨运行累积会与随机 IP 撞车 → flaky
        await test_session.execute(delete(Device).where(Device.ip_address.in_([ip_a, ip_b])))
        await test_session.commit()


@pytest.mark.asyncio
async def test_devices_import_rejects_bad_type(client):
    """非 xlsx/csv 文件导入 → 40001。"""
    manager_t = await _login(client, "manager01")
    resp = await client.post("/api/v1/monitor/devices/import", headers=_h(manager_t),
                             files={"file": ("evil.txt", b"not excel", "text/plain")})
    assert resp.json()["code"] == 40001
    assert "xlsx" in resp.json()["message"]


@pytest.mark.asyncio
async def test_file_upload_whitelist(client, monkeypatch):
    """上传白名单：可执行/脚本扩展名拒绝；MIME 明显不匹配拒绝；白名单内正常上传。"""
    from app.api.v1 import files as files_mod

    class _FakeBucket:
        def __init__(self):
            self.objects = []

        def bucket_exists(self, bucket):
            return True

        def make_bucket(self, bucket):
            return None

        def put_object(self, bucket, key, data, length, content_type=""):
            self.objects.append((bucket, key, length))
            return None

        def presigned_get_object(self, bucket, key, expires):
            return f"http://minio.local/{key}"

    fake = _FakeBucket()
    monkeypatch.setattr(files_mod, "_client", lambda: fake)
    manager_t = await _login(client, "manager01")
    h = _h(manager_t)

    # .exe → 拒绝
    resp = await client.post("/api/v1/files", headers=h, files={"file": ("malware.exe", b"x", "application/x-msdownload")})
    assert resp.json()["code"] == 40001
    assert "不支持" in resp.json()["message"]

    # .png 但 MIME 是脚本 → 拒绝（改扩展名伪装）
    resp = await client.post("/api/v1/files", headers=h,
                             files={"file": ("evil.png", b"\x89PNG\r\n\x1a\n", "application/x-sh")})
    assert resp.json()["code"] == 40001
    assert "不匹配" in resp.json()["message"]

    # 白名单内 .png + image/png → 成功，存入 MinIO 替身
    resp = await client.post("/api/v1/files", headers=h,
                             files={"file": ("shot.png", b"\x89PNG\r\n\x1a\n", "image/png")})
    assert resp.json()["code"] == 0, resp.json()
    assert fake.objects and fake.objects[0][2] == 8


# ---------- 设备监控优化：数据范围补全 / 权限收紧 / 告警去重 / 巡检 ----------
@pytest.mark.asyncio
async def test_alert_list_dept_data_scope(client, test_session):
    """告警列表数据范围：dept 角色仅见本部门设备的告警，无设备全局告警不可见。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    db = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()   # 攻防实验室
    web = (await test_session.execute(select(Device).where(Device.name == "web-01"))).scalar_one()  # 安全运营部

    tag = uuid.uuid4().hex[:6]
    r = await client.post("/api/v1/monitor/alerts", headers=_h(analyst_t), json={
        "title": f"dept-alert-{tag}", "severity": "high", "alert_type": "abnormal", "device_id": db.id,
        "description": "scoped",
    })
    assert r.json()["code"] == 0, r.json()
    await client.post("/api/v1/monitor/alerts", headers=_h(manager_t), json={
        "title": f"other-alert-{tag}", "severity": "high", "alert_type": "abnormal", "device_id": web.id,
        "description": "other dept",
    })
    await client.post("/api/v1/monitor/alerts", headers=_h(manager_t), json={
        "title": f"global-alert-{tag}", "severity": "low", "alert_type": "abnormal", "description": "no device",
    })

    items = (await client.get("/api/v1/monitor/alerts", headers=_h(analyst_t), params={"size": 100})).json()["data"]["items"]
    titles = {i["title"] for i in items}
    assert f"dept-alert-{tag}" in titles
    assert f"other-alert-{tag}" not in titles
    assert f"global-alert-{tag}" not in titles

    m_items = (await client.get("/api/v1/monitor/alerts", headers=_h(manager_t), params={"size": 100})).json()["data"]["items"]
    m_titles = {i["title"] for i in m_items}
    assert {f"dept-alert-{tag}", f"other-alert-{tag}", f"global-alert-{tag}"} <= m_titles


@pytest.mark.asyncio
async def test_scan_report_list_dept_data_scope(client, test_session):
    """扫描报告列表数据范围：dept 角色仅见本部门设备的报告。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    db = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()   # 攻防实验室
    web = (await test_session.execute(select(Device).where(Device.name == "web-01"))).scalar_one()  # 安全运营部
    test_session.add_all([
        ScanReport(target_ip="10.0.10.12", device_id=db.id, report_type="on_demand", scan_status="completed",
                   generated_by=1, status="pending_review"),
        ScanReport(target_ip="10.0.10.11", device_id=web.id, report_type="on_demand", scan_status="completed",
                   generated_by=1, status="pending_review"),
        ScanReport(target_ip="10.99.99.9", device_id=None, report_type="on_demand", scan_status="completed",
                   generated_by=1, status="pending_review"),
    ])
    await test_session.commit()

    analyst_targets = {i["target_ip"] for i in (await client.get("/api/v1/monitor/scans/reports", headers=_h(analyst_t), params={"size": 100})).json()["data"]["items"]}
    assert "10.0.10.12" in analyst_targets       # db-01（本部门）
    assert "10.0.10.11" not in analyst_targets   # web-01（其他部门）
    assert "10.99.99.9" not in analyst_targets   # 无设备全局报告

    m_targets = {i["target_ip"] for i in (await client.get("/api/v1/monitor/scans/reports", headers=_h(manager_t), params={"size": 100})).json()["data"]["items"]}
    assert {"10.0.10.12", "10.0.10.11", "10.99.99.9"} <= m_targets


@pytest.mark.asyncio
async def test_auditor_ping_forbidden(client, test_session):
    """探测会改写设备状态 → 权限收紧为 manage；auditor（仅 view）→ 40302。"""
    auditor_t = await _login(client, "auditor01")
    db = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()

    resp = await client.post(f"/api/v1/monitor/devices/{db.id}/ping", headers=_h(auditor_t))
    assert resp.json()["code"] == 40302

    # analyst（有 device:manage）仍可正常探测
    analyst_t = await _login(client, "analyst01")
    resp = await client.post(f"/api/v1/monitor/devices/{db.id}/ping", headers=_h(analyst_t))
    assert resp.json()["code"] == 0, resp.json()
    assert resp.json()["data"]["last_seen_at"]


@pytest.mark.asyncio
async def test_scan_alert_dedup_window(client, test_session, monkeypatch):
    """自动告警去重：同 target 高危扫描连续跑两次只产生一条告警，且外部通知只触发一次。"""
    from app.services import scanner as scanner_mod

    async def _high_nmap(target, ports, svc, scan_options=None):
        return 0, _HIGH_RISK_XML, ""

    notified = []

    async def _fake_notify(alert_id, title, content, severity):
        notified.append(alert_id)

    monkeypatch.setattr("app.services.scanner._run_nmap", _high_nmap)
    monkeypatch.setattr("app.services.scanner.notify_alert_task", _fake_notify)

    # 双段随机目标：10.99.{x}.11 仅 200 个取值，跨运行残留的扫描告警会撞随机目标 → dedup 短路 → flaky。
    # 双段全随机（约 6.4 万取值）+ conftest 会话启动清理扫描告警，消除跨运行/测试间碰撞
    target = f"10.{int(uuid.uuid4().hex[:4], 16) % 254 + 1}.{int(uuid.uuid4().hex[:4], 16) % 254 + 1}.11"
    for _ in range(2):
        r = ScanReport(target_ip=target, report_type="on_demand", scan_status="pending", generated_by=1)
        test_session.add(r)
        await test_session.commit()
        await scanner_mod.execute_scan(r.id, target, 100)

    rows = (await test_session.execute(select(Alert).where(Alert.title.like(f"%{target}%")))).scalars().all()
    assert len(rows) == 1
    # INET 列读回为 ipaddress 对象，归一化比较
    assert str(rows[0].target_ip) == target and rows[0].alert_type == "intrusion"
    assert len(notified) == 1  # 去重后只触发一次外部通知


@pytest.mark.asyncio
async def test_device_list_includes_offline_since(client, test_session):
    """设备列表响应带 offline_since（巡检判定离线时间），供前端展示离线自。"""
    manager_t = await _login(client, "manager01")
    db = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()
    db.offline_since = dt.datetime.now(dt.timezone.utc)
    await test_session.commit()

    resp = await client.get("/api/v1/monitor/devices", headers=_h(manager_t), params={"size": 100})
    item = next(i for i in resp.json()["data"]["items"] if i["id"] == db.id)
    assert item["offline_since"] is not None


@pytest.mark.asyncio
async def test_patrols_list_data_scope(client, test_session):
    """巡检历史列表数据范围：dept 角色仅见本部门子网的巡检记录（含分组计数）。"""
    manager_t = await _login(client, "manager01")
    analyst_t = await _login(client, "analyst01")

    biz = (await test_session.execute(select(IPSubnet).where(IPSubnet.network == "10.0.10.0/24"))).scalar_one()   # 攻防实验室
    office = (await test_session.execute(select(IPSubnet).where(IPSubnet.network == "10.0.0.0/24"))).scalar_one()  # 安全运营部
    now = dt.datetime.now(dt.timezone.utc)
    p1 = DevicePatrol(subnet_id=biz.id, network="10.0.10.0/24", scan_status="completed",
                      online_ips=["10.0.10.11"], offline_ips=[], ghost_ips=[], started_at=now, completed_at=now)
    p2 = DevicePatrol(subnet_id=office.id, network="10.0.0.0/24", scan_status="completed",
                      online_ips=[], offline_ips=[], ghost_ips=[], started_at=now, completed_at=now)
    test_session.add_all([p1, p2])
    await test_session.commit()

    a_items = (await client.get("/api/v1/monitor/patrols", headers=_h(analyst_t), params={"size": 50})).json()["data"]["items"]
    a_ids = {i["id"] for i in a_items}
    assert p1.id in a_ids and p2.id not in a_ids

    m_items = (await client.get("/api/v1/monitor/patrols", headers=_h(manager_t), params={"size": 50})).json()["data"]["items"]
    m_ids = {i["id"] for i in m_items}
    assert p1.id in m_ids and p2.id in m_ids
    item = next(i for i in m_items if i["id"] == p1.id)
    assert item["online_count"] == 1 and item["ghost_count"] == 0 and item["offline_count"] == 0
    assert item["subnet_name"] == "业务网"


@pytest.mark.asyncio
async def test_delete_device_archives_when_referenced(client, test_session):
    """删除保护：设备被扫描报告引用时归档保留（不再因外键裸 500）。"""
    analyst_t = await _login(client, "analyst01")
    ip = _uniq_ip()
    resp = await client.post("/api/v1/monitor/devices", headers=_h(analyst_t), json={
        "name": "ref-scanned", "ip_address": ip, "device_type": "server", "status": "active",
    })
    assert resp.json()["code"] == 0, resp.json()
    dev_id = resp.json()["data"]["id"]

    report = ScanReport(device_id=dev_id, target_ip=ip, report_type="on_demand",
                        scan_status="failed", status="pending_review", generated_by=1, error="x")
    test_session.add(report)
    await test_session.commit()
    try:
        resp = await client.request("DELETE", f"/api/v1/monitor/devices/{dev_id}", headers=_h(analyst_t), params={"reason": "清理"})
        assert resp.json()["code"] == 0, resp.json()
        assert "已归档保留" in resp.json()["data"]["message"]
        d = await test_session.get(Device, dev_id)
        assert d.status == "archived"
    finally:
        await test_session.execute(delete(ScanReport).where(ScanReport.device_id == dev_id))
        await test_session.execute(delete(Device).where(Device.id == dev_id))
        await test_session.commit()


@pytest.mark.asyncio
async def test_get_device_data_scope(client, test_session):
    """越权防护：dept 范围角色不能读他部门设备详情（与列表端点一致）。"""
    analyst_t = await _login(client, "analyst01")
    web = (await test_session.execute(select(Device).where(Device.name == "web-01"))).scalar_one()
    db = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()

    resp = await client.get(f"/api/v1/monitor/devices/{web.id}", headers=_h(analyst_t))  # web-01 属安全运营部
    assert resp.json()["code"] == 40301, resp.json()
    resp = await client.get(f"/api/v1/monitor/devices/{db.id}", headers=_h(analyst_t))  # db-01 属攻防实验室
    assert resp.json()["code"] == 0, resp.json()


@pytest.mark.asyncio
async def test_get_scan_report_scope_and_creator(client, test_session):
    """扫描报告数据范围：dept 角色不能读他部门设备报告；创建者始终可读自己的无设备报告。"""
    from app.models import User

    analyst_t = await _login(client, "analyst01")
    manager_t = await _login(client, "manager01")
    analyst = (await test_session.execute(select(User).where(User.username == "analyst01"))).scalar_one()
    manager = (await test_session.execute(select(User).where(User.username == "manager01"))).scalar_one()
    web = (await test_session.execute(select(Device).where(Device.name == "web-01"))).scalar_one()

    other = ScanReport(device_id=web.id, target_ip="10.0.10.11", report_type="on_demand",
                       scan_status="completed", status="pending_review", generated_by=manager.id)
    mine = ScanReport(device_id=None, target_ip="10.0.10.99", report_type="on_demand",
                      scan_status="completed", status="pending_review", generated_by=analyst.id)
    test_session.add_all([other, mine])
    await test_session.commit()
    try:
        # dept 角色读他部门设备（web-01 → 安全运营部）的报告 → 403
        resp = await client.get(f"/api/v1/monitor/scans/reports/{other.id}", headers=_h(analyst_t))
        assert resp.json()["code"] == 40301, resp.json()
        # 创建者（manager，all 范围）可读
        resp = await client.get(f"/api/v1/monitor/scans/reports/{other.id}", headers=_h(manager_t))
        assert resp.json()["code"] == 0, resp.json()
        # 无设备报告：创建者（analyst）可读，他人（manager）也可读（all 范围）
        resp = await client.get(f"/api/v1/monitor/scans/reports/{mine.id}", headers=_h(analyst_t))
        assert resp.json()["code"] == 0, resp.json()
    finally:
        await test_session.execute(delete(ScanReport).where(ScanReport.id.in_([other.id, mine.id])))
        await test_session.commit()


# ---------- 批次3：输入校验与裸 500 治理 ----------
@pytest.mark.asyncio
async def test_device_mac_format_validation(client, test_session):
    """设备 MAC 格式校验（批次3）：非法 MAC 创建/更新 → 40001。"""
    manager_t = await _login(client, "manager01")
    ip = _uniq_ip()
    resp = await client.post("/api/v1/monitor/devices", headers=_h(manager_t), json={
        "name": "bad-mac", "ip_address": ip, "mac_address": "ZZ:ZZ:00:11:22", "device_type": "server",
    })
    assert resp.json()["code"] == 40001, resp.json()
    assert "MAC 地址格式" in resp.json()["message"]

    # 合法 MAC 创建成功，随后更新为非法 → 40001
    good_mac = f"02:{uuid.uuid4().hex[:2].upper()}:{uuid.uuid4().hex[:2].upper()}:AA:BB:CC"
    resp = await client.post("/api/v1/monitor/devices", headers=_h(manager_t), json={
        "name": "good-mac", "ip_address": ip, "mac_address": good_mac, "device_type": "server",
    })
    assert resp.json()["code"] == 0, resp.json()
    dev_id = resp.json()["data"]["id"]
    try:
        resp = await client.put(f"/api/v1/monitor/devices/{dev_id}", headers=_h(manager_t),
                                json={"mac_address": "nope"})
        assert resp.json()["code"] == 40001, resp.json()
    finally:
        await test_session.execute(delete(Device).where(Device.id == dev_id))
        await test_session.commit()


@pytest.mark.asyncio
async def test_subnet_gateway_format_validation(client):
    """子网网关格式校验（批次3）：非法 gateway → 40001。"""
    manager_t = await _login(client, "manager01")
    net = f"10.210.{int(uuid.uuid4().hex[:4], 16) % 240 + 1}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={
        "name": "bad-gw", "network": net, "gateway": "not-an-ip",
    })
    assert resp.json()["code"] == 40001, resp.json()
    assert "网关格式不正确" in resp.json()["message"]


@pytest.mark.asyncio
async def test_allocation_fk_validation(client, test_session):
    """IP 分配外键校验（批次3）：allocated_to / device_id 不存在 → 40400。"""
    manager_t = await _login(client, "manager01")
    net = f"10.211.{int(uuid.uuid4().hex[:4], 16) % 240 + 1}.0/24"
    resp = await client.post("/api/v1/monitor/subnets", headers=_h(manager_t), json={"name": "fk-net", "network": net})
    assert resp.json()["code"] == 0, resp.json()
    subnet_id = resp.json()["data"]["id"]
    try:
        resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
            "subnet_id": subnet_id, "allocated_to": 999999, "allocation_type": "static", "purpose": "坏用户",
        })
        assert resp.json()["code"] == 40400, resp.json()
        assert "分配用户不存在" in resp.json()["message"]

        resp = await client.post("/api/v1/monitor/allocations", headers=_h(manager_t), json={
            "subnet_id": subnet_id, "device_id": 999999, "allocation_type": "static", "purpose": "坏设备",
        })
        assert resp.json()["code"] == 40400, resp.json()
        assert "绑定设备不存在" in resp.json()["message"]
    finally:
        await test_session.execute(delete(IPSubnet).where(IPSubnet.id == subnet_id))
        await test_session.commit()


@pytest.mark.asyncio
async def test_file_upload_size_precheck(client, monkeypatch):
    """文件上传内存治理（批次3）：Content-Length 超限预检返回 40001，不读入内存。"""
    from app.api.v1 import files as files_mod

    monkeypatch.setattr(files_mod.settings, "UPLOAD_MAX_SIZE_MB", 1)
    manager_t = await _login(client, "manager01")
    resp = await client.post("/api/v1/files", headers=_h(manager_t),
                             files={"file": ("big.png", b"\x89PNG\r\n\x1a\n" + b"\x00" * (2 * 1024 * 1024), "image/png")})
    assert resp.json()["code"] == 40001, resp.json()
    assert "超过 1MB 限制" in resp.json()["message"]


# ---------- 批次4：设备写操作数据范围 ----------
@pytest.mark.asyncio
async def test_device_write_scope(client):
    """设备写操作数据范围（批次4）：dept 角色不能改/删/探测他部门设备，可管理自己创建的设备（自动归属本部门）。"""
    analyst_t = await _login(client, "analyst01")
    manager_t = await _login(client, "manager01")

    web = next(
        d for d in (await client.get("/api/v1/monitor/devices", headers=_h(manager_t), params={"size": 50})).json()["data"]["items"]
        if d["name"] == "web-01"
    )
    # analyst 更新/探测/删除他部门（web-01 → 安全运营部）设备 → 403
    resp = await client.put(f"/api/v1/monitor/devices/{web['id']}", headers=_h(analyst_t), json={"location": "x"})
    assert resp.json()["code"] == 40301, resp.json()
    resp = await client.post(f"/api/v1/monitor/devices/{web['id']}/ping", headers=_h(analyst_t))
    assert resp.json()["code"] == 40301, resp.json()
    resp = await client.request("DELETE", f"/api/v1/monitor/devices/{web['id']}", headers=_h(analyst_t), params={"reason": "x"})
    assert resp.json()["code"] == 40301, resp.json()

    # 自己创建的设备（未指定部门 → 自动归属创建者部门）可更新/探测/删除
    ip = _uniq_ip()
    resp = await client.post("/api/v1/monitor/devices", headers=_h(analyst_t), json={
        "name": "scope-srv", "ip_address": ip, "device_type": "server",
    })
    assert resp.json()["code"] == 0, resp.json()
    dev_id = resp.json()["data"]["id"]
    assert resp.json()["data"]["department_id"] is not None  # 归属 analyst 所在部门
    resp = await client.put(f"/api/v1/monitor/devices/{dev_id}", headers=_h(analyst_t), json={"location": "机房A"})
    assert resp.json()["code"] == 0, resp.json()
    resp = await client.post(f"/api/v1/monitor/devices/{dev_id}/ping", headers=_h(analyst_t))
    assert resp.json()["code"] == 0, resp.json()
    resp = await client.request("DELETE", f"/api/v1/monitor/devices/{dev_id}", headers=_h(analyst_t), params={"reason": "清理"})
    assert resp.json()["code"] == 0, resp.json()


# ==================== 扫描增强：scan_options / NSE / 取消重试 / 基线漂移 ====================

@pytest.mark.asyncio
async def test_scan_with_scan_options(client, test_session, monkeypatch):
    """scan_type/port_range/nse 传入 → scan_options 落库，列表与详情可见。"""
    analyst_t = await _login(client, "analyst01")
    monkeypatch.setattr("app.services.scanner._run_nmap", _fake_nmap)

    # dept 角色列表按设备部门过滤：关联本部门设备 db-01 保证列表可见
    db_dev = (await test_session.execute(select(Device).where(Device.name == "db-01"))).scalar_one()
    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={
        "target_ip": _scan_ip(), "scan_type": "sT", "port_range": "22,80,443", "nse": False,
        "device_id": db_dev.id,
    })
    assert resp.json()["code"] == 0, resp.json()
    report_id = resp.json()["data"]["report_id"]

    detail = None
    for _ in range(200):
        resp = await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))
        detail = resp.json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "completed", detail
    assert detail["scan_options"]["scan_type"] == "sT"
    assert detail["scan_options"]["port_range"] == "22,80,443"
    assert detail["scan_options"]["nse"] is False
    assert detail["error_code"] is None

    item = (await client.get("/api/v1/monitor/scans/reports", headers=_h(analyst_t))).json()["data"]["items"]
    row = next(r for r in item if r["id"] == report_id)
    assert row["scan_options"]["scan_type"] == "sT" and "error_code" in row


@pytest.mark.asyncio
async def test_scan_invalid_port_range_rejected(client, test_session):
    """非法端口范围（区间越界）与 ports/port_range 互斥 → 业务校验拒绝。"""
    analyst_t = await _login(client, "analyst01")
    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={
        "target_ip": _scan_ip(), "port_range": "0-99999",
    })
    assert resp.json()["code"] == 40001
    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={
        "target_ip": _scan_ip(), "ports": 100, "port_range": "22,80",
    })
    assert resp.json()["code"] == 40001


@pytest.mark.asyncio
async def test_scan_nse_detection(client, test_session, monkeypatch):
    """NSE 脚本结果 → vulnerabilities 出现 source=nse 条目（真实 CVE 检测）。"""
    analyst_t = await _login(client, "analyst01")
    NSE_XML = """<?xml version="1.0"?>
<nmaprun scanner="nmap" version="7.94">
  <host><status state="up"/>
    <ports>
      <port protocol="tcp" portid="445">
        <state state="open"/>
        <service name="microsoft-ds"/>
        <script id="smb-vuln-ms17-010" output="VULNERABLE: MS17-010 remote code execution"/>
      </port>
    </ports>
  </host>
</nmaprun>
"""

    async def _nse_nmap(target, ports, svc, scan_options=None):
        return 0, NSE_XML, ""

    monkeypatch.setattr("app.services.scanner._run_nmap", _nse_nmap)

    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": _scan_ip()})
    assert resp.json()["code"] == 0, resp.json()
    report_id = resp.json()["data"]["report_id"]

    detail = None
    for _ in range(200):
        resp = await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))
        detail = resp.json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "completed", detail
    assert detail["scan_data"]["nse_scripts"] == "vuln"  # NMAP_NSE_SCRIPTS 默认值
    nse_vulns = [v for v in detail["scan_data"]["vulnerabilities"] if v["source"] == "nse"]
    assert any(v["name"] == "smb-vuln-ms17-010" and v["severity"] == "critical" for v in nse_vulns)


@pytest.mark.asyncio
async def test_scan_cancel(client, test_session, monkeypatch):
    """运行中扫描可取消 → 落 failed/cancelled；已完成/失败任务不可再取消。"""
    analyst_t = await _login(client, "analyst01")

    async def _slow_nmap_cancel(target, ports, svc, scan_options=None):
        await asyncio.sleep(30)
        return 0, SAMPLE_XML, ""

    monkeypatch.setattr("app.services.scanner._run_nmap", _slow_nmap_cancel)

    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": _scan_ip()})
    report_id = resp.json()["data"]["report_id"]

    running = False
    for _ in range(200):
        detail = (await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))).json()["data"]
        if detail["scan_status"] == "running":
            running = True
            break
        await asyncio.sleep(0.05)
    assert running, "扫描未进入 running 状态"

    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/cancel", headers=_h(analyst_t))
    assert resp.json()["code"] == 0, resp.json()

    detail = None
    for _ in range(200):
        detail = (await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))).json()["data"]
        if detail["scan_status"] == "failed":
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "failed"
    assert detail["error_code"] == "cancelled"

    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/cancel", headers=_h(analyst_t))
    assert resp.json()["code"] == 40001


@pytest.mark.asyncio
async def test_scan_retry(client, test_session, monkeypatch):
    """失败任务重试 → 恢复 pending 重新执行，沿用原 scan_options，错误被清空。"""
    analyst_t = await _login(client, "analyst01")

    async def _boom(target, ports, svc, scan_options=None):
        raise RuntimeError("boom")

    monkeypatch.setattr("app.services.scanner._run_nmap", _boom)

    resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": _scan_ip()})
    report_id = resp.json()["data"]["report_id"]

    for _ in range(200):
        detail = (await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))).json()["data"]
        if detail["scan_status"] == "failed":
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "failed" and detail["error_code"] == "generic"

    monkeypatch.setattr("app.services.scanner._run_nmap", _fake_nmap)
    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/retry", headers=_h(analyst_t))
    assert resp.json()["code"] == 0, resp.json()

    for _ in range(200):
        detail = (await client.get(f"/api/v1/monitor/scans/reports/{report_id}", headers=_h(analyst_t))).json()["data"]
        if detail["scan_status"] in ("completed", "failed"):
            break
        await asyncio.sleep(0.05)
    assert detail["scan_status"] == "completed", detail
    assert detail["error"] is None and detail["error_code"] is None
    assert detail["scan_options"]["scan_type"] == "sS"  # 沿用原选项

    resp = await client.post(f"/api/v1/monitor/scans/reports/{report_id}/retry", headers=_h(analyst_t))
    assert resp.json()["code"] == 40001


@pytest.mark.asyncio
async def test_scan_baseline_diff(client, test_session, monkeypatch):
    """同目标两次扫描 → 第二次 scan_data.baseline_diff 对比出新增/关闭端口。"""
    analyst_t = await _login(client, "analyst01")
    target = _scan_ip()
    BASE2_XML = """<?xml version="1.0"?>
<nmaprun scanner="nmap" version="7.94">
  <host><status state="up"/>
    <ports>
      <port protocol="tcp" portid="22"><state state="open"/>
        <service name="ssh" product="OpenSSH" version="8.9p1"/></port>
      <port protocol="tcp" portid="80"><state state="open"/>
        <service name="http"/></port>
    </ports>
  </host>
</nmaprun>
"""
    xmls = [SAMPLE_XML, BASE2_XML]

    async def _seq_nmap(target, ports, svc, scan_options=None):
        return 0, xmls.pop(0), ""

    monkeypatch.setattr("app.services.scanner._run_nmap", _seq_nmap)

    ids = []
    for _ in range(2):
        resp = await client.post("/api/v1/monitor/scans", headers=_h(analyst_t), json={"target_ip": target})
        assert resp.json()["code"] == 0, resp.json()
        ids.append(resp.json()["data"]["report_id"])

    details = []
    for rid in ids:
        for _ in range(200):
            d = (await client.get(f"/api/v1/monitor/scans/reports/{rid}", headers=_h(analyst_t))).json()["data"]
            if d["scan_status"] in ("completed", "failed"):
                break
            await asyncio.sleep(0.05)
        details.append(d)
    assert details[0]["scan_status"] == details[1]["scan_status"] == "completed"

    # 首次扫描无基线；第二次与首次对比（SAMPLE_XML 含 22/6379，BASE2_XML 含 22/80）
    assert details[0]["scan_data"]["baseline_diff"] is None
    diff = details[1]["scan_data"]["baseline_diff"]
    assert {p["port"] for p in diff["new_ports"]} == {80}
    assert {p["port"] for p in diff["closed_ports"]} == {6379}
    assert diff["changed_services"] == []
    assert "与上次扫描相比" in details[1]["summary"]
